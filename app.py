#!/usr/bin/env python3

from __future__ import annotations

import base64
import csv
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import tarfile
import tempfile
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from pathlib import Path
from zoneinfo import ZoneInfo

import requests as http_requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
try:
    import weasyprint
except Exception:  # pragma: no cover - optional runtime dependency
    weasyprint = None

from flask import Flask, abort, after_this_request, flash, g, jsonify, make_response, redirect, render_template, request, send_file, send_from_directory, session, url_for
from markupsafe import Markup
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
# Allow database path to be set via environment variable for Docker support
DEFAULT_DATABASE_FILENAME = "statement_software.db"
LEGACY_DATABASE_FILENAME = os.environ.get("LEGACY_DATABASE_FILENAME", "").strip()
DB_PATH = Path(os.environ.get("DATABASE_PATH", BASE_DIR / DEFAULT_DATABASE_FILENAME))
DATA_DIR = DB_PATH.parent
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", DATA_DIR / "uploads"))
BACKUP_DIR = Path(os.environ.get("BACKUP_DIR", DATA_DIR / "backups"))
try:
    MAX_UPLOAD_MB = max(1, int(os.environ.get("MAX_UPLOAD_MB", "512")))
except ValueError:
    MAX_UPLOAD_MB = 512
SEED_DEMO_DATA = os.environ.get("SEED_DEMO_DATA", "0").strip().lower() in {"1", "true", "yes", "on"}
SOURCE_CSV_PATH = os.environ.get("SOURCE_CSV_PATH", "").strip()
SOURCE_CSV = Path(SOURCE_CSV_PATH) if SOURCE_CSV_PATH else None
DEMO_CLIENT_NAME = os.environ.get("DEMO_CLIENT_NAME", "Demo Client").strip() or "Demo Client"
UNCATEGORIZED = "uncategorized"
BACKUP_FORMAT_VERSION = 1
FULL_BACKUP_NOTES = "Statement Software full backup. Includes SQLite database and uploads only."
APP_NAME = os.environ.get("APP_NAME", "Statement Software").strip() or "Statement Software"
BRAND_NAME = os.environ.get("BRAND_NAME", "Statement").strip() or "Statement"
COMPANY_NAME = os.environ.get("COMPANY_NAME", "Your Company").strip() or "Your Company"
CHINA_TZ = ZoneInfo("Asia/Shanghai")
ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".heic"}
ALLOWED_DB_EXT = {".db", ".sqlite", ".sqlite3"}
IMAGE_MIME_MAP = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp", ".heic": "image/heic",
}
OPENROUTER_API_KEY_ENV = os.environ.get("OPENROUTER_API_KEY", "")
OPENROUTER_MODELS_URL = "https://openrouter.ai/api/v1/models"
DEFAULT_OPENROUTER_MODEL = "google/gemini-2.5-flash"
RESET_SECRET_TOKEN = os.environ.get("RESET_SECRET_TOKEN", "")
SECRET_KEY_FROM_ENV = bool(os.environ.get("SECRET_KEY"))
INITIAL_ADMIN_USERNAME = os.environ.get("INITIAL_ADMIN_USERNAME", "").strip()
INITIAL_ADMIN_PASSWORD = os.environ.get("INITIAL_ADMIN_PASSWORD", "")
INITIAL_ADMIN_MUST_CHANGE = os.environ.get("INITIAL_ADMIN_MUST_CHANGE", "0") == "1"
BOOTSTRAP_CREDENTIAL_PATH = DATA_DIR / "admin_bootstrap.txt"
DEFAULT_PROFIT_EXPENSE_ACCOUNT_NAME = (
    os.environ.get("DEFAULT_PROFIT_EXPENSE_ACCOUNT_NAME", "Company Profit").strip()
    or "Company Profit"
)


def _asset_version() -> str:
    explicit = os.environ.get("ASSET_VERSION", "").strip()
    if explicit:
        return explicit
    candidates = [
        BASE_DIR / "app.py",
        BASE_DIR / "static" / "styles.css",
        BASE_DIR / "static" / "mobile.css",
        BASE_DIR / "templates" / "_calculator.html",
    ]
    newest = max((path.stat().st_mtime_ns for path in candidates if path.exists()), default=0)
    return str(newest)


ASSET_VERSION = _asset_version()
REQUIRED_DB_TABLES = {
    "users",
    "clients",
    "statement_entries",
    "app_settings",
    "api_tokens",
    "quick_submits",
    "expense_accounts",
    "expense_entries",
}
SAFE_HTTP_METHODS = {"GET", "HEAD", "OPTIONS", "TRACE"}

# Cached vision-capable models from OpenRouter
_openrouter_models_cache: dict = {}  # {"models": [...], "fetched_at": float}
_MODELS_CACHE_TTL = 3600  # 1 hour


def _fetch_openrouter_models() -> list[tuple[str, str]]:
    """Fetch vision-capable models from OpenRouter API, with 1-hour cache."""
    import time
    now = time.time()
    cached = _openrouter_models_cache
    if cached and (now - cached["fetched_at"]) < _MODELS_CACHE_TTL:
        return cached["models"]

    try:
        resp = http_requests.get(OPENROUTER_MODELS_URL, timeout=10)
        resp.raise_for_status()
        data = resp.json().get("data", [])
        models = []
        for m in data:
            input_mods = m.get("architecture", {}).get("input_modalities", [])
            if "image" in input_mods:
                model_id = m["id"]
                name = m.get("name", model_id)
                ctx = m.get("context_length")
                label = f"{name} ({ctx // 1000}K)" if ctx else name
                models.append((model_id, label))
        _openrouter_models_cache["models"] = models
        _openrouter_models_cache["fetched_at"] = now
        return models
    except Exception:
        # Return cached if available, otherwise a minimal fallback
        if cached and cached.get("models"):
            return cached["models"]
        return [(DEFAULT_OPENROUTER_MODEL, "Gemini 2.5 Flash (default)")]

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.secret_key = os.environ.get("SECRET_KEY", "")
app.config["PERMANENT_SESSION_LIFETIME"] = 8 * 60 * 60  # 8 hours
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "0") == "1"


def ensure_runtime_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def refresh_app_secret_key(db: sqlite3.Connection | None = None) -> None:
    if SECRET_KEY_FROM_ENV:
        app.secret_key = os.environ["SECRET_KEY"]
        return
    owns_connection = db is None
    conn = db or sqlite3.connect(DB_PATH)
    if owns_connection:
        conn.row_factory = sqlite3.Row
    existing_key = conn.execute("select value from app_settings where key = 'secret_key'").fetchone()
    if existing_key:
        app.secret_key = existing_key["value"] if isinstance(existing_key, sqlite3.Row) else existing_key[0]
    else:
        new_key = secrets.token_hex(32)
        conn.execute("insert or replace into app_settings(key, value) values ('secret_key', ?)", (new_key,))
        app.secret_key = new_key
        if owns_connection:
            conn.commit()
    if owns_connection:
        conn.close()


def get_csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def csrf_input() -> Markup:
    token = get_csrf_token()
    return Markup(f'<input type="hidden" name="csrf_token" value="{token}">')


@app.context_processor
def inject_template_helpers() -> dict:
    return {
        "csrf_token": get_csrf_token,
        "csrf_input": csrf_input,
        "app_name": APP_NAME,
        "brand_name": BRAND_NAME,
        "company_name": COMPANY_NAME,
        "default_profit_expense_account_name": DEFAULT_PROFIT_EXPENSE_ACCOUNT_NAME,
        "asset_version": ASSET_VERSION,
    }


def get_setting(key: str, default: str = "") -> str:
    db = get_db()
    row = db.execute("select value from app_settings where key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    db = get_db()
    db.execute("insert or replace into app_settings(key, value) values (?, ?)", (key, value))
    db.commit()


def get_openrouter_api_key() -> str:
    return get_setting("openrouter_api_key", OPENROUTER_API_KEY_ENV)


def get_openrouter_model() -> str:
    return get_setting("openrouter_model", DEFAULT_OPENROUTER_MODEL)


def _delete_image_file(image_path: str | None) -> None:
    if image_path:
        (UPLOAD_DIR / image_path).unlink(missing_ok=True)


def _call_vision(image_data: str, mime: str, system_prompt: str, user_prompt: str = "Extract the transaction details from this image.") -> str:
    from openrouter import OpenRouter
    api_key = get_openrouter_api_key()
    model = get_openrouter_model()
    if not api_key:
        raise ValueError("OpenRouter API key not configured. Go to Settings to add it.")
    with OpenRouter(api_key=api_key) as client:
        response = client.chat.send(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{image_data}"}},
                        {"type": "text", "text": user_prompt},
                    ],
                },
            ],
            temperature=0.1,
        )
    return response.choices[0].message.content


def _resolve_image_upload(file=None, image_url=None):
    """Resolve image from file upload or URL to (image_data_b64, mime). Raises ValueError on error."""
    if file:
        fname = file.filename or "image.jpg"
        ext = os.path.splitext(fname)[1].lower() or ".jpg"
        if ext not in ALLOWED_IMAGE_EXT:
            raise ValueError("Invalid image type")
        image_data = base64.b64encode(file.read()).decode("utf-8")
        mime = file.content_type if file.content_type and file.content_type.startswith("image/") else IMAGE_MIME_MAP.get(ext, "image/jpeg")
        return image_data, mime
    elif image_url:
        if image_url.startswith("data:"):
            m = re.match(r"data:(image/\w+);base64,(.+)", image_url)
            if not m:
                raise ValueError("Invalid data URL")
            return m.group(2), m.group(1)
        elif "/uploads/" in image_url:
            filename = os.path.basename(image_url.split("/uploads/")[-1])
            filepath = UPLOAD_DIR / filename
            if not filepath.resolve().is_relative_to(UPLOAD_DIR.resolve()):
                raise ValueError("Invalid image path")
            if not filepath.exists():
                raise ValueError("Image not found")
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ALLOWED_IMAGE_EXT:
                raise ValueError("Invalid image type")
            with open(filepath, "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8"), IMAGE_MIME_MAP.get(ext, "image/jpeg")
        raise ValueError("Invalid image URL")
    raise ValueError("No image provided")


def currency_symbol(code: str) -> str:
    return {"USD": "$", "CNY": "¥", "EGP": "E£"}.get(code, "")


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def utc_timestamp() -> str:
    return utc_now().isoformat(timespec="seconds").replace("+00:00", "Z")


def render_pdf_bytes(html_string: str) -> bytes:
    if weasyprint is None:
        raise RuntimeError("PDF export requires WeasyPrint to be installed in the runtime environment.")
    return weasyprint.HTML(string=html_string).write_pdf()


def china_today() -> date:
    return utc_now().astimezone(CHINA_TZ).date()


def normalize_utc_timestamp(value: str | None) -> str:
    if not value:
        return ""
    if value.endswith("Z"):
        return value
    if "T" in value:
        return f"{value}Z"
    return value


@app.template_filter("money")
def money_filter(value: float) -> str:
    return f"{float(value):,.2f}"


@app.template_filter("money_with_symbol")
def money_with_symbol_filter(value_currency: tuple[float, str]) -> str:
    value, code = value_currency
    return f"{currency_symbol(code)}{float(value):,.2f}"


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_: object) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def close_active_db() -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def _slugify_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-_.")


def format_file_size(num_bytes: int) -> str:
    value = float(num_bytes)
    units = ["B", "KB", "MB", "GB"]
    for unit in units:
        if value < 1024 or unit == units[-1]:
            precision = 0 if unit == "B" else 1
            return f"{value:.{precision}f} {unit}"
        value /= 1024
    return f"{num_bytes} B"


def database_file_info(path: Path) -> dict:
    stat = path.stat()
    created_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
    return {
        "name": path.name,
        "size_bytes": stat.st_size,
        "size_label": format_file_size(stat.st_size),
        "created_at": created_at.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "kind": "Full backup" if is_full_backup_file(path) else "Database only",
    }


def is_full_backup_file(path: Path) -> bool:
    name = path.name.lower()
    return name.endswith(".tar.gz") or name.endswith(".tgz")


def is_restore_file(path: Path) -> bool:
    return path.suffix.lower() in ALLOWED_DB_EXT or is_full_backup_file(path)


def list_database_backups() -> list[dict]:
    ensure_runtime_dirs()
    paths = [
        path for path in BACKUP_DIR.iterdir()
        if path.is_file() and is_restore_file(path)
    ]
    paths.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return [database_file_info(path) for path in paths]


def resolve_backup_path(name: str) -> Path:
    safe_name = Path(name).name
    if safe_name != name:
        abort(404)
    path = BACKUP_DIR / safe_name
    if not path.exists() or not is_restore_file(path):
        abort(404)
    return path


def make_database_backup_name(label: str) -> str:
    timestamp = utc_now().strftime("%Y%m%dT%H%M%SZ")
    suffix = _slugify_filename(label) or "backup"
    return f"{timestamp}-{suffix}-{uuid.uuid4().hex[:8]}.db"


def make_full_backup_name(label: str = "") -> str:
    timestamp = utc_now().strftime("%Y%m%d-%H%M%S")
    suffix = _slugify_filename(label)
    label_part = f"-{suffix}" if suffix else ""
    return f"statement-full-backup-{timestamp}{label_part}.tar.gz"


def snapshot_database(target_path: Path) -> Path:
    ensure_runtime_dirs()
    copy_database_contents(DB_PATH, target_path)
    return target_path


def copy_database_contents(source_path: Path, target_path: Path) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    source = sqlite3.connect(str(source_path))
    destination = sqlite3.connect(str(target_path))
    try:
        with destination:
            source.backup(destination)
    finally:
        destination.close()
        source.close()
    return target_path


def create_database_backup(label: str) -> Path:
    backup_path = BACKUP_DIR / make_database_backup_name(label)
    return snapshot_database(backup_path)


def count_upload_files(path: Path) -> int:
    if not path.exists():
        return 0
    return sum(1 for item in path.rglob("*") if item.is_file())


def create_full_backup(target_path: Path) -> Path:
    ensure_runtime_dirs()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=DATA_DIR) as tmp:
        tmp_path = Path(tmp)
        db_copy = tmp_path / DEFAULT_DATABASE_FILENAME
        manifest_path = tmp_path / "manifest.json"
        uploads_source = UPLOAD_DIR
        if not uploads_source.exists():
            uploads_source = tmp_path / "uploads"
            uploads_source.mkdir()
        snapshot_database(db_copy)
        manifest = {
            "format_version": BACKUP_FORMAT_VERSION,
            "app_name": APP_NAME,
            "created_at_utc": utc_timestamp(),
            "database_file": DEFAULT_DATABASE_FILENAME,
            "uploads_dir": "uploads",
            "upload_file_count": count_upload_files(uploads_source),
            "notes": FULL_BACKUP_NOTES,
        }
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        with tarfile.open(target_path, "w:gz") as tar:
            tar.add(manifest_path, arcname="manifest.json")
            tar.add(db_copy, arcname=DEFAULT_DATABASE_FILENAME)
            tar.add(uploads_source, arcname="uploads")
    return target_path


def create_full_backup_file(label: str = "") -> Path:
    return create_full_backup(BACKUP_DIR / make_full_backup_name(label))


def validate_database_file(path: Path) -> None:
    db = None
    tables: set[str] = set()
    try:
        db = sqlite3.connect(str(path))
        integrity = db.execute("pragma integrity_check").fetchone()
        if not integrity or str(integrity[0]).lower() != "ok":
            raise ValueError("Database integrity check failed.")
        tables = {
            row[0] for row in db.execute(
                "select name from sqlite_master where type = 'table'"
            ).fetchall()
        }
    except sqlite3.DatabaseError as exc:
        raise ValueError("The uploaded file is not a valid SQLite database.") from exc
    finally:
        if db is not None:
            db.close()
    missing = sorted(REQUIRED_DB_TABLES - tables)
    if missing:
        raise ValueError("Database is missing required tables: " + ", ".join(missing))


def validate_backup_manifest(path: Path) -> None:
    if not path.exists():
        return
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise ValueError("Backup manifest is not valid JSON.") from exc
    if manifest.get("format_version") != BACKUP_FORMAT_VERSION:
        raise ValueError("Backup manifest format is not supported.")


def extract_backup_archive(archive_path: Path, target_dir: Path) -> None:
    try:
        with tarfile.open(archive_path, "r:gz") as tar:
            base = target_dir.resolve()
            for member in tar.getmembers():
                if member.issym() or member.islnk():
                    raise ValueError(f"Backup archive contains unsafe link: {member.name}")
                target = (target_dir / member.name).resolve()
                if not target.is_relative_to(base):
                    raise ValueError(f"Backup archive contains unsafe path: {member.name}")
            tar.extractall(target_dir)
    except tarfile.TarError as exc:
        raise ValueError("Backup archive is not a valid .tar.gz file.") from exc


def stage_restore_source(source_path: Path, original_name: str) -> tuple[Path, Path | None, str, str, Path]:
    ensure_runtime_dirs()
    staging_dir = Path(tempfile.mkdtemp(prefix="restore-", dir=DATA_DIR))
    candidate_db = staging_dir / DEFAULT_DATABASE_FILENAME
    staged_uploads: Path | None = None
    try:
        if is_full_backup_file(source_path):
            extracted_dir = staging_dir / "archive"
            extracted_dir.mkdir()
            extract_backup_archive(source_path, extracted_dir)
            validate_backup_manifest(extracted_dir / "manifest.json")
            manifest_path = extracted_dir / "manifest.json"
            restored_db = extracted_dir / DEFAULT_DATABASE_FILENAME
            if manifest_path.exists():
                try:
                    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                    manifest_database = str(manifest.get("database_file") or "").strip()
                    if manifest_database:
                        restored_db = extracted_dir / manifest_database
                except (json.JSONDecodeError, OSError):
                    pass
            if not restored_db.exists() and LEGACY_DATABASE_FILENAME:
                restored_db = extracted_dir / LEGACY_DATABASE_FILENAME
            if not restored_db.exists():
                raise ValueError(f"Backup archive does not contain {DEFAULT_DATABASE_FILENAME}.")
            validate_database_file(restored_db)
            copy_database_contents(restored_db, candidate_db)
            staged_uploads = extracted_dir / "uploads"
            staged_uploads.mkdir(exist_ok=True)
            return candidate_db, staged_uploads, original_name, "full", staging_dir

        if source_path.suffix.lower() in ALLOWED_DB_EXT:
            shutil.copy2(source_path, candidate_db)
            validate_database_file(candidate_db)
            return candidate_db, None, original_name, "database", staging_dir

        raise ValueError("Restore file must be a full .tar.gz/.tgz backup or a .db/.sqlite/.sqlite3 database.")
    except Exception:
        shutil.rmtree(staging_dir, ignore_errors=True)
        raise


def stage_uploaded_restore(file_storage) -> tuple[Path, Path | None, str, str, Path]:
    if not file_storage or not file_storage.filename:
        raise ValueError("Choose a full backup or database file to upload.")
    original_name = Path(file_storage.filename).name
    lower_name = original_name.lower()
    if not (
        lower_name.endswith(".tar.gz")
        or lower_name.endswith(".tgz")
        or Path(original_name).suffix.lower() in ALLOWED_DB_EXT
    ):
        raise ValueError("Restore file must end in .tar.gz, .tgz, .db, .sqlite, or .sqlite3.")
    ensure_runtime_dirs()
    suffix = ".tar.gz" if lower_name.endswith(".tar.gz") else Path(original_name).suffix
    with tempfile.NamedTemporaryFile(delete=False, dir=DATA_DIR, suffix=suffix) as tmp:
        uploaded_path = Path(tmp.name)
    try:
        file_storage.save(uploaded_path)
        return stage_restore_source(uploaded_path, original_name)
    finally:
        uploaded_path.unlink(missing_ok=True)


def stage_backup_restore(name: str) -> tuple[Path, Path | None, str, str, Path]:
    backup_path = resolve_backup_path(name)
    return stage_restore_source(backup_path, backup_path.name)


def clear_directory_contents(directory: Path) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    for child in directory.iterdir():
        if child.is_dir() and not child.is_symlink():
            shutil.rmtree(child)
        else:
            child.unlink(missing_ok=True)


def replace_uploads_contents(source_dir: Path | None) -> None:
    clear_directory_contents(UPLOAD_DIR)
    if source_dir is None or not source_dir.exists():
        return
    for child in source_dir.iterdir():
        target = UPLOAD_DIR / child.name
        if child.is_dir() and not child.is_symlink():
            shutil.copytree(child, target)
        else:
            shutil.copy2(child, target)


def activate_restore_candidate(candidate_path: Path, label: str, staged_uploads: Path | None = None) -> Path:
    validate_database_file(candidate_path)
    backup_path = create_full_backup_file(f"before-{label}")
    close_active_db()
    with tempfile.TemporaryDirectory(dir=DATA_DIR) as tmp:
        tmp_path = Path(tmp)
        rollback_db = tmp_path / "rollback.db"
        rollback_uploads = tmp_path / "uploads"
        snapshot_database(rollback_db)
        if staged_uploads is not None and UPLOAD_DIR.exists():
            shutil.copytree(UPLOAD_DIR, rollback_uploads)
        try:
            copy_database_contents(candidate_path, DB_PATH)
            if staged_uploads is not None:
                replace_uploads_contents(staged_uploads)
            init_db()
        except Exception:
            copy_database_contents(rollback_db, DB_PATH)
            if staged_uploads is not None:
                replace_uploads_contents(rollback_uploads if rollback_uploads.exists() else None)
            init_db()
            raise
    return backup_path


def activate_database_candidate(candidate_path: Path, label: str) -> Path:
    return activate_restore_candidate(candidate_path, label)


def create_admin_bootstrap_file(password: str) -> None:
    ensure_runtime_dirs()
    content = (
        f"{APP_NAME} bootstrap admin credentials\n"
        f"Generated: {utc_timestamp()}\n"
        "Username: admin\n"
        f"Password: {password}\n"
        "Change this password immediately after first login.\n"
    )
    BOOTSTRAP_CREDENTIAL_PATH.write_text(content, encoding="utf-8")
    try:
        os.chmod(BOOTSTRAP_CREDENTIAL_PATH, 0o600)
    except OSError:
        pass
    print(f"[bootstrap] Admin credentials written to {BOOTSTRAP_CREDENTIAL_PATH}")


# --- Auth helpers ---

PUBLIC_ENDPOINTS = {"login_page", "login_action", "static"}
API_ENDPOINTS = set()  # populated by @api_route decorator


def _validate_api_token() -> bool:
    """Check Authorization header or X-API-Key for a valid token. Returns True if valid."""
    auth = request.headers.get("Authorization", "")
    api_key = request.headers.get("X-API-Key", "")
    raw_token = ""
    if auth.startswith("Bearer "):
        raw_token = auth[7:].strip()
    elif api_key:
        raw_token = api_key.strip()
    if not raw_token:
        return False
    from hashlib import sha256
    token_hash = sha256(raw_token.encode()).hexdigest()
    db = get_db()
    row = db.execute("select * from api_tokens where token_hash = ? and is_active = 1", (token_hash,)).fetchone()
    if row:
        db.execute("update api_tokens set last_used_at = ? where id = ?", (utc_timestamp(), row["id"]))
        db.commit()
        return True
    return False


def _api_log(action: str, resource_type: str, resource_id: int | None, detail: dict, undo_data: dict | None = None) -> None:
    """Record an API write action to the audit log."""
    db = get_db()
    db.execute(
        "insert into api_audit_log (action, resource_type, resource_id, detail, undo_data, created_at) values (?, ?, ?, ?, ?, ?)",
        (action, resource_type, resource_id, json.dumps(detail), json.dumps(undo_data) if undo_data else None, utc_timestamp()),
    )
    db.commit()


def api_route(rule, **options):
    """Decorator that registers a route as an API endpoint (token-auth, bypasses session)."""
    def decorator(f):
        endpoint_name = options.pop("endpoint", f.__name__)
        API_ENDPOINTS.add(endpoint_name)
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not _validate_api_token():
                return jsonify({"error": "Invalid or missing API token"}), 401
            return f(*args, **kwargs)
        app.add_url_rule(rule, endpoint=endpoint_name, view_func=wrapper, **options)
        return wrapper
    return decorator


def _is_ajax_request() -> bool:
    return (
        request.headers.get("X-Requested-With") == "XMLHttpRequest"
        or bool(request.headers.get("X-CSRF-Token"))
        or request.is_json
    )


def _csrf_error_response():
    message = "CSRF token missing or invalid."
    if _is_ajax_request():
        return jsonify({"error": message}), 400
    abort(400, description=message)


@app.before_request
def require_login():
    g.user = None
    endpoint = request.endpoint
    if endpoint in PUBLIC_ENDPOINTS or endpoint in API_ENDPOINTS or endpoint is None:
        return
    if "user_id" not in session:
        if request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login_page"))
    db = get_db()
    user = db.execute("select * from users where id = ?", (session["user_id"],)).fetchone()
    if not user or not user["is_active"]:
        session.clear()
        return redirect(url_for("login_page"))
    g.user = user
    # Force password change if needed (except on change-password and logout pages)
    if user["must_change_password"] and endpoint not in ("change_password", "logout"):
        return redirect(url_for("change_password"))


@app.before_request
def verify_csrf_token():
    endpoint = request.endpoint
    if request.method in SAFE_HTTP_METHODS or endpoint in API_ENDPOINTS or endpoint is None:
        return
    expected = session.get("_csrf_token")
    provided = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not expected or not provided or not secrets.compare_digest(expected, provided):
        return _csrf_error_response()


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not g.user or g.user["role"] != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated


def init_db() -> None:
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.executescript(
        """
        create table if not exists clients (
            id integer primary key autoincrement,
            name text not null unique
        );

        create table if not exists statement_entries (
            id integer primary key autoincrement,
            client_id integer not null,
            source_no integer not null,
            entry_date text not null,
            description text not null,
            currency text not null,
            direction text not null,
            amount real not null,
            kind text not null,
            category_hint text not null,
            transfer_group text,
            exchange_rate real,
            linked_entry_id integer,
            commission_source_entry_id integer,
            profit_expense_entry_id integer,
            profit_expense_account_id integer,
            foreign key(client_id) references clients(id),
            foreign key(commission_source_entry_id) references statement_entries(id),
            foreign key(profit_expense_entry_id) references expense_entries(id),
            foreign key(profit_expense_account_id) references expense_accounts(id)
        );

        create table if not exists bank_balances (
            id integer primary key autoincrement,
            account_name text not null,
            usd_balance real not null default 0,
            cny_balance real not null default 0,
            updated_at text not null
        );

        create table if not exists supplier_balances (
            id integer primary key autoincrement,
            supplier_name text not null,
            currency text not null default 'CNY',
            amount_owed real not null default 0,
            notes text not null default '',
            updated_at text not null
        );

        create table if not exists statement_entry_events (
            id integer primary key autoincrement,
            client_id integer not null,
            entry_id integer,
            action text not null,
            payload text not null,
            created_at text not null,
            undone_at text,
            foreign key(client_id) references clients(id)
        );

        create table if not exists app_settings (
            key text primary key,
            value text not null
        );

        create table if not exists users (
            id integer primary key autoincrement,
            username text not null unique,
            password_hash text not null,
            role text not null default 'user',
            is_active integer not null default 1,
            must_change_password integer not null default 0,
            created_at text not null default '',
            last_login text
        );

        create table if not exists expense_accounts (
            id integer primary key autoincrement,
            name text not null unique,
            enabled_currencies text not null default 'CNY',
            created_at text not null default ''
        );

        create table if not exists expense_entries (
            id integer primary key autoincrement,
            account_id integer not null,
            seq_no integer not null default 0,
            entry_date text not null,
            description text not null,
            currency text not null,
            direction text not null default 'OUT',
            amount real not null,
            category text not null default 'general',
            is_recurring integer not null default 0,
            template_id integer,
            image_path text,
            linked_statement_entry_id integer,
            created_at text not null default '',
            foreign key(account_id) references expense_accounts(id),
            foreign key(linked_statement_entry_id) references statement_entries(id)
        );

        create table if not exists recurring_expense_templates (
            id integer primary key autoincrement,
            account_id integer not null,
            description text not null,
            currency text not null default 'CNY',
            direction text not null default 'OUT',
            amount real not null,
            day_of_month integer not null default 1,
            category text not null default 'general',
            is_active integer not null default 1,
            last_generated text not null default '',
            created_at text not null default '',
            foreign key(account_id) references expense_accounts(id)
        );

        create table if not exists expense_events (
            id integer primary key autoincrement,
            account_id integer not null,
            entry_id integer,
            action text not null,
            payload text not null,
            created_at text not null,
            undone_at text,
            foreign key(account_id) references expense_accounts(id)
        );

        create table if not exists quick_submits (
            id integer primary key autoincrement,
            client_id integer not null,
            description text not null default '',
            amount real,
            image_path text not null,
            status text not null default 'pending',
            created_at text not null,
            processed_at text,
            created_entry_id integer,
            created_by text not null default '',
            foreign key(client_id) references clients(id)
        );

        create table if not exists api_tokens (
            id integer primary key autoincrement,
            name text not null,
            token_hash text not null unique,
            token_prefix text not null,
            created_at text not null,
            last_used_at text,
            is_active integer not null default 1
        );

        create table if not exists api_audit_log (
            id integer primary key autoincrement,
            action text not null,
            resource_type text not null,
            resource_id integer,
            detail text not null default '{}',
            undo_data text,
            undone integer not null default 0,
            created_at text not null
        );
        """
    )
    # Seed bootstrap admin if no users exist
    user_count = db.execute("select count(*) from users").fetchone()[0]
    if user_count == 0:
        if INITIAL_ADMIN_USERNAME and INITIAL_ADMIN_PASSWORD:
            db.execute(
                "insert into users(username, password_hash, role, must_change_password, created_at) values (?, ?, 'admin', ?, ?)",
                (
                    INITIAL_ADMIN_USERNAME,
                    generate_password_hash(INITIAL_ADMIN_PASSWORD),
                    1 if INITIAL_ADMIN_MUST_CHANGE else 0,
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
            print(f"[bootstrap] Admin user '{INITIAL_ADMIN_USERNAME}' created from installer configuration.")
        else:
            bootstrap_password = secrets.token_urlsafe(12)
            db.execute(
                "insert into users(username, password_hash, role, must_change_password, created_at) values (?, ?, 'admin', 1, ?)",
                ("admin", generate_password_hash(bootstrap_password), datetime.now(timezone.utc).isoformat()),
            )
            create_admin_bootstrap_file(bootstrap_password)
    # Generate SECRET_KEY and store in DB if not set via env
    refresh_app_secret_key(db)
    entry_columns = {row[1] for row in db.execute("pragma table_info(statement_entries)").fetchall()}
    if "exchange_rate" not in entry_columns:
        db.execute("alter table statement_entries add column exchange_rate real")
    if "linked_entry_id" not in entry_columns:
        db.execute("alter table statement_entries add column linked_entry_id integer")
    if "image_path" not in entry_columns:
        db.execute("alter table statement_entries add column image_path text")
    if "commission_source_entry_id" not in entry_columns:
        db.execute("alter table statement_entries add column commission_source_entry_id integer")
    if "profit_expense_entry_id" not in entry_columns:
        db.execute("alter table statement_entries add column profit_expense_entry_id integer")
    if "profit_expense_account_id" not in entry_columns:
        db.execute("alter table statement_entries add column profit_expense_account_id integer")
    db.execute(
        "create index if not exists idx_statement_entries_commission_source on statement_entries(commission_source_entry_id)"
    )
    db.execute(
        "create index if not exists idx_statement_entries_profit_expense_entry on statement_entries(profit_expense_entry_id)"
    )
    db.execute(
        "create index if not exists idx_statement_entries_profit_expense_account on statement_entries(profit_expense_account_id)"
    )
    client_columns = {row[1] for row in db.execute("pragma table_info(clients)").fetchall()}
    if "parent_id" not in client_columns:
        db.execute("alter table clients add column parent_id integer references clients(id)")
    expense_columns = {row[1] for row in db.execute("pragma table_info(expense_entries)").fetchall()}
    if "linked_statement_entry_id" not in expense_columns:
        db.execute("alter table expense_entries add column linked_statement_entry_id integer")
    db.execute(
        "create index if not exists idx_expense_entries_linked_statement on expense_entries(linked_statement_entry_id)"
    )
    qs_columns = {row[1] for row in db.execute("pragma table_info(quick_submits)").fetchall()}
    if qs_columns and "amount" not in qs_columns:
        db.execute("alter table quick_submits add column amount real")
    tpl_columns = {row[1] for row in db.execute("pragma table_info(recurring_expense_templates)").fetchall()}
    if "every_n_months" not in tpl_columns:
        db.execute("alter table recurring_expense_templates add column every_n_months integer not null default 1")
    backfill_exchange_links(db)
    db.commit()
    db.close()
    ensure_runtime_dirs()


def seed_from_csv() -> bool:
    if SOURCE_CSV is None:
        print("[seed] SOURCE_CSV_PATH is not set; skipping demo import.")
        return False
    if not SOURCE_CSV.exists():
        print(f"[seed] Demo CSV not found at {SOURCE_CSV}; skipping demo import.")
        return False

    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()
    cur.execute("insert or ignore into clients(name) values (?)", (DEMO_CLIENT_NAME,))
    client_id = cur.execute("select id from clients where name = ?", (DEMO_CLIENT_NAME,)).fetchone()[0]

    cur.execute("delete from statement_entries where client_id = ?", (client_id,))

    with SOURCE_CSV.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cur.execute(
                """
                insert into statement_entries (
                    client_id, source_no, entry_date, description, currency, direction,
                    amount, kind, category_hint, transfer_group
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client_id,
                    int(float(row["source_no"])),
                    row["date"],
                    row["description"],
                    row["currency"],
                    row["direction"],
                    float(row["amount"]),
                    row["kind"],
                    UNCATEGORIZED,
                    row["transfer_group"] or None,
                ),
            )

    db.commit()
    db.row_factory = sqlite3.Row
    entry_ids = db.execute(
        f"select id from statement_entries where client_id = ? order by {ENTRY_ORDER}",
        (client_id,),
    ).fetchall()
    for index, row in enumerate(entry_ids, start=1):
        db.execute("update statement_entries set source_no = ? where id = ?", (index, row["id"]))
    db.commit()
    db.close()
    return True


def ensure_seeded() -> None:
    db = sqlite3.connect(DB_PATH)
    count = db.execute("select count(*) from statement_entries").fetchone()[0]
    db.close()
    if count > 0:
        return
    if not SEED_DEMO_DATA:
        print("[seed] Demo data disabled; starting with an empty database.")
        return
    seed_from_csv()


def client_row_data(db: sqlite3.Connection, client_id: int, name: str, entry_count: int, parent_id: int | None = None) -> dict:
    usd_in = db.execute(
        "select coalesce(sum(amount),0) from statement_entries where client_id=? and currency='USD' and direction='IN'",
        (client_id,),
    ).fetchone()[0]
    usd_out = db.execute(
        "select coalesce(sum(amount),0) from statement_entries where client_id=? and currency='USD' and direction='OUT'",
        (client_id,),
    ).fetchone()[0]
    cny_in = db.execute(
        "select coalesce(sum(amount),0) from statement_entries where client_id=? and currency='CNY' and direction='IN'",
        (client_id,),
    ).fetchone()[0]
    cny_out = db.execute(
        "select coalesce(sum(amount),0) from statement_entries where client_id=? and currency='CNY' and direction='OUT'",
        (client_id,),
    ).fetchone()[0]
    last_date = db.execute(
        "select max(entry_date) from statement_entries where client_id=?",
        (client_id,),
    ).fetchone()[0]
    return {
        "id": client_id,
        "name": name,
        "entry_count": entry_count,
        "parent_id": parent_id,
        "usd_balance": usd_in - usd_out,
        "cny_balance": cny_in - cny_out,
        "last_date": last_date or "",
    }


def client_list() -> list[dict]:
    db = get_db()
    rows = db.execute(
        """
        select c.id, c.name, c.parent_id, count(se.id) as entry_count
        from clients c
        left join statement_entries se on se.client_id = c.id
        group by c.id, c.name, c.parent_id
        order by c.name
        """
    ).fetchall()
    clients = []
    for row in rows:
        clients.append(client_row_data(db, row["id"], row["name"], row["entry_count"], row["parent_id"]))
    return clients


def grouped_client_list() -> list[dict]:
    """Return clients organized into groups.

    A 'group' is a top-level dict with:
      - If the client has children: the parent info + 'children' list
      - If standalone (no parent, no children): just the client + empty children list

    Children are clients whose parent_id points to a top-level client.
    """
    all_clients = client_list()
    by_id = {c["id"]: c for c in all_clients}
    parent_ids = {c["parent_id"] for c in all_clients if c["parent_id"]}

    groups = []
    used = set()

    # First, build groups for parents that have children
    for pid in sorted(parent_ids):
        if pid not in by_id:
            continue
        parent = dict(by_id[pid])
        children = [c for c in all_clients if c["parent_id"] == pid]
        # Aggregate totals across parent + children
        all_in_group = [parent] + children
        parent["group_usd"] = sum(c["usd_balance"] for c in all_in_group)
        parent["group_cny"] = sum(c["cny_balance"] for c in all_in_group)
        parent["group_entries"] = sum(c["entry_count"] for c in all_in_group)
        parent["children"] = children
        groups.append(parent)
        used.add(pid)
        for c in children:
            used.add(c["id"])

    # Then add standalone clients (no parent, not a parent of anything)
    for c in all_clients:
        if c["id"] not in used:
            standalone = dict(c)
            standalone["children"] = []
            standalone["group_usd"] = c["usd_balance"]
            standalone["group_cny"] = c["cny_balance"]
            standalone["group_entries"] = c["entry_count"]
            groups.append(standalone)

    groups.sort(key=lambda g: g["name"].lower())
    return groups


def dashboard_stats() -> dict:
    db = get_db()
    total_entries = db.execute("select count(*) from statement_entries").fetchone()[0]
    total_clients = db.execute("select count(*) from clients").fetchone()[0]
    total_usd_in = db.execute("select coalesce(sum(amount),0) from statement_entries where currency='USD' and direction='IN'").fetchone()[0]
    total_usd_out = db.execute("select coalesce(sum(amount),0) from statement_entries where currency='USD' and direction='OUT'").fetchone()[0]
    total_cny_in = db.execute("select coalesce(sum(amount),0) from statement_entries where currency='CNY' and direction='IN'").fetchone()[0]
    total_cny_out = db.execute("select coalesce(sum(amount),0) from statement_entries where currency='CNY' and direction='OUT'").fetchone()[0]
    return {
        "total_entries": total_entries,
        "total_clients": total_clients,
        "total_usd_balance": total_usd_in - total_usd_out,
        "total_cny_balance": total_cny_in - total_cny_out,
        "total_usd_in": total_usd_in,
        "total_usd_out": total_usd_out,
        "total_cny_in": total_cny_in,
        "total_cny_out": total_cny_out,
    }


def bank_balance_list() -> list[dict]:
    db = get_db()
    rows = db.execute("select id, account_name, usd_balance, cny_balance, updated_at from bank_balances order by account_name").fetchall()
    return [dict(r) for r in rows]


def bank_balance_totals(balances: list[dict]) -> dict:
    return {
        "total_usd": sum(b["usd_balance"] for b in balances),
        "total_cny": sum(b["cny_balance"] for b in balances),
    }


def supplier_balance_list() -> list[dict]:
    db = get_db()
    rows = db.execute("select id, supplier_name, currency, amount_owed, notes, updated_at from supplier_balances order by supplier_name").fetchall()
    return [dict(r) for r in rows]


def supplier_balance_totals(suppliers: list[dict]) -> dict:
    return {
        "total_usd": sum(s["amount_owed"] for s in suppliers if s["currency"] == "USD"),
        "total_cny": sum(s["amount_owed"] for s in suppliers if s["currency"] == "CNY"),
    }


def sync_exchange_group(db: sqlite3.Connection, transfer_group: str | None) -> None:
    if not transfer_group:
        return
    group_entries = db.execute(
        """
        select id, client_id, currency, direction, amount
        from statement_entries
        where transfer_group = ?
        order by id
        """,
        (transfer_group,),
    ).fetchall()
    if not group_entries:
        return

    usd_out = [row for row in group_entries if row["currency"] == "USD" and row["direction"] == "OUT"]
    cny_in = [row for row in group_entries if row["currency"] == "CNY" and row["direction"] == "IN"]
    usd_amount = sum(float(row["amount"]) for row in usd_out)
    cny_amount = sum(float(row["amount"]) for row in cny_in)
    exchange_rate = (cny_amount / usd_amount) if usd_amount > 0 and cny_amount > 0 else None

    for row in group_entries:
        db.execute(
            "update statement_entries set exchange_rate = ?, linked_entry_id = null where id = ?",
            (exchange_rate, row["id"]),
        )

    if len(usd_out) == len(cny_in):
        for usd_row, cny_row in zip(usd_out, cny_in):
            db.execute("update statement_entries set linked_entry_id = ? where id = ?", (cny_row["id"], usd_row["id"]))
            db.execute("update statement_entries set linked_entry_id = ? where id = ?", (usd_row["id"], cny_row["id"]))


def backfill_exchange_links(db: sqlite3.Connection) -> None:
    groups = db.execute(
        "select distinct transfer_group from statement_entries where transfer_group is not null and trim(transfer_group) != ''"
    ).fetchall()
    for row in groups:
        sync_exchange_group(db, row[0])


def commission_child_entry(db: sqlite3.Connection, source_entry_id: int) -> sqlite3.Row | None:
    return db.execute(
        """
        select *
        from statement_entries
        where commission_source_entry_id = ?
        order by id
        limit 1
        """,
        (source_entry_id,),
    ).fetchone()


def commission_relationship_maps(
    db: sqlite3.Connection, client_id: int
) -> tuple[dict[int, dict[str, int | None]], dict[int, dict[str, int | None]]]:
    rows = db.execute(
        """
        select
            child.id as child_id,
            child.source_no as child_source_no,
            child.commission_source_entry_id as source_entry_id,
            source.source_no as source_source_no
        from statement_entries child
        left join statement_entries source on source.id = child.commission_source_entry_id
        where child.client_id = ? and child.commission_source_entry_id is not null
        """,
        (client_id,),
    ).fetchall()
    child_by_source: dict[int, dict[str, int | None]] = {}
    source_by_child: dict[int, dict[str, int | None]] = {}
    for row in rows:
        source_entry_id = row["source_entry_id"]
        child_id = row["child_id"]
        if source_entry_id is None:
            continue
        child_by_source[source_entry_id] = {
            "id": child_id,
            "source_no": row["child_source_no"],
        }
        source_by_child[child_id] = {
            "id": source_entry_id,
            "source_no": row["source_source_no"],
        }
    return child_by_source, source_by_child


def annotate_commission_rows(
    rows: list[dict], child_by_source: dict[int, dict[str, int | None]], source_by_child: dict[int, dict[str, int | None]]
) -> list[dict]:
    annotated: list[dict] = []
    for row in rows:
        row_data = dict(row)
        child = child_by_source.get(row_data["id"])
        source = source_by_child.get(row_data["id"])
        row_data["commission_child_id"] = child["id"] if child else None
        row_data["commission_child_source_no"] = child["source_no"] if child else None
        row_data["commission_source_source_no"] = source["source_no"] if source else None
        row_data["can_auto_commission"] = (
            row_data.get("direction") == "OUT"
            and not row_data.get("commission_source_entry_id")
            and not child
        )
        annotated.append(row_data)
    return annotated


def statement_rows_with_commission_state(
    db: sqlite3.Connection, client_id: int, entries: list[sqlite3.Row]
) -> list[dict]:
    rows = running_balances(entries)
    child_by_source, source_by_child = commission_relationship_maps(db, client_id)
    rows = annotate_commission_rows(rows, child_by_source, source_by_child)
    profit_account_ids = {
        int(row["profit_expense_account_id"])
        for row in rows
        if row.get("profit_expense_account_id")
    }
    account_names: dict[int, str] = {}
    if profit_account_ids:
        placeholders = ",".join("?" for _ in profit_account_ids)
        account_rows = db.execute(
            f"select id, name from expense_accounts where id in ({placeholders})",
            tuple(profit_account_ids),
        ).fetchall()
        account_names = {int(row["id"]): row["name"] for row in account_rows}
    for row in rows:
        account_id = row.get("profit_expense_account_id")
        row["profit_expense_account_name"] = account_names.get(int(account_id)) if account_id else None
    return rows


def commission_error_message(
    db: sqlite3.Connection, source_entry: sqlite3.Row | None, client_id: int
) -> str | None:
    if source_entry is None:
        return "Source entry not found."
    if source_entry["client_id"] != client_id:
        return "Source entry does not belong to this client."
    if source_entry["commission_source_entry_id"]:
        return "Auto commission cannot start from a generated commission row."
    if source_entry["direction"] != "OUT":
        return "Auto commission is only available for outgoing entries."
    if commission_child_entry(db, source_entry["id"]) is not None:
        return "This entry already has a generated commission."
    return None


def parse_commission_percentage(raw_value: object) -> Decimal:
    try:
        percentage = Decimal(str(raw_value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Commission percentage must be a valid number.") from exc
    if percentage <= 0:
        raise ValueError("Commission percentage must be greater than 0.")
    return percentage


def calculate_commission_amount(source_amount: object, percentage: Decimal) -> float:
    amount = (
        Decimal(str(source_amount)) * percentage / Decimal("100")
    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if amount <= 0:
        raise ValueError("Commission amount must be greater than 0 after rounding.")
    return float(amount)


def commission_description(source_description: str, note: str | None) -> str:
    clean_note = (note or "").strip()
    if clean_note:
        return clean_note
    clean_source = (source_description or "").strip()
    return f"Commission - {clean_source}" if clean_source else "Commission"


def create_commission_entry(
    db: sqlite3.Connection,
    source_entry: sqlite3.Row,
    entry_date: str,
    percentage_raw: object,
    note: str | None = None,
) -> tuple[sqlite3.Row, float]:
    percentage = parse_commission_percentage(percentage_raw)
    amount = calculate_commission_amount(source_entry["amount"], percentage)
    description = commission_description(source_entry["description"], note)
    cursor = db.execute(
        """
        insert into statement_entries (
            client_id, source_no, entry_date, description, currency, direction,
            amount, kind, category_hint, commission_source_entry_id
        ) values (?, 0, ?, ?, ?, 'OUT', ?, 'movement', 'commission', ?)
        """,
        (
            source_entry["client_id"],
            entry_date,
            description,
            source_entry["currency"],
            amount,
            source_entry["id"],
        ),
    )
    created_id = cursor.lastrowid
    resequence_client_entries(source_entry["client_id"])
    created = db.execute("select * from statement_entries where id = ?", (created_id,)).fetchone()
    if created is None:
        raise ValueError("Commission entry could not be created.")
    record_event(
        source_entry["client_id"],
        created_id,
        "commission",
        {
            "percentage": float(percentage),
            "source_entry": row_to_dict(source_entry),
            "entry": row_to_dict(created),
        },
    )
    return created, float(percentage)


def parse_bool_flag(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def parse_optional_int(value: object, field_label: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_label} must be a valid integer.") from exc


def expense_account_currencies(account: sqlite3.Row | dict | None) -> list[str]:
    if not account:
        return []
    enabled = account["enabled_currencies"] if isinstance(account, sqlite3.Row) else account.get("enabled_currencies", "")
    return [c.strip() for c in str(enabled).split(",") if c.strip()]


def expense_account_options(db: sqlite3.Connection) -> list[dict]:
    rows = db.execute(
        "select id, name, enabled_currencies from expense_accounts order by lower(name), id"
    ).fetchall()
    options: list[dict] = []
    for row in rows:
        currencies = expense_account_currencies(row)
        options.append(
            {
                "id": row["id"],
                "name": row["name"],
                "enabled_currencies": row["enabled_currencies"],
                "currencies": currencies,
                "currency_label": ", ".join(currencies),
            }
        )
    return options


def default_profit_expense_account_id(options: list[dict]) -> int | None:
    for option in options:
        if option["name"].strip().lower() == DEFAULT_PROFIT_EXPENSE_ACCOUNT_NAME.lower():
            return int(option["id"])
    return None


def validate_profit_expense_selection(
    db: sqlite3.Connection,
    *,
    enabled: bool,
    account_id: int | None,
    category_hint: str,
    currency: str,
) -> sqlite3.Row | None:
    if not enabled:
        return None
    if category_hint != "commission":
        raise ValueError("Company profit can only be enabled for commission entries.")
    if not account_id:
        raise ValueError("Choose an expense account for company profit.")
    account = db.execute(
        "select * from expense_accounts where id = ?",
        (account_id,),
    ).fetchone()
    if account is None:
        raise ValueError("Selected expense account no longer exists.")
    if currency not in expense_account_currencies(account):
        raise ValueError(f"Expense account '{account['name']}' does not support {currency}.")
    return account


def linked_profit_expense_entry(
    db: sqlite3.Connection,
    statement_entry_id: int,
    profit_expense_entry_id: int | None = None,
) -> sqlite3.Row | None:
    row = None
    if profit_expense_entry_id:
        row = db.execute(
            "select * from expense_entries where id = ?",
            (profit_expense_entry_id,),
        ).fetchone()
        if row is not None:
            return row
    return db.execute(
        """
        select *
        from expense_entries
        where linked_statement_entry_id = ?
        order by id
        limit 1
        """,
        (statement_entry_id,),
    ).fetchone()


def sync_statement_profit_entry(
    db: sqlite3.Connection,
    statement_entry_id: int,
    *,
    enabled: bool,
    account_id: int | None,
) -> set[int]:
    affected_accounts: set[int] = set()
    statement_entry = db.execute(
        "select * from statement_entries where id = ?",
        (statement_entry_id,),
    ).fetchone()
    if statement_entry is None:
        return affected_accounts

    existing_expense = linked_profit_expense_entry(
        db,
        statement_entry_id,
        statement_entry["profit_expense_entry_id"] if "profit_expense_entry_id" in statement_entry.keys() else None,
    )

    if not enabled or statement_entry["category_hint"] != "commission":
        if existing_expense is not None:
            affected_accounts.add(existing_expense["account_id"])
            _delete_image_file(existing_expense["image_path"] if "image_path" in existing_expense.keys() else None)
            db.execute("delete from expense_entries where id = ?", (existing_expense["id"],))
        db.execute(
            "update statement_entries set profit_expense_entry_id = null, profit_expense_account_id = null where id = ?",
            (statement_entry_id,),
        )
        return affected_accounts

    account = validate_profit_expense_selection(
        db,
        enabled=enabled,
        account_id=account_id,
        category_hint=statement_entry["category_hint"],
        currency=statement_entry["currency"],
    )
    assert account is not None
    description = statement_entry["description"].strip() or "Commission"

    if existing_expense is None:
        cursor = db.execute(
            """
            insert into expense_entries (
                account_id, seq_no, entry_date, description, currency, direction,
                amount, category, is_recurring, template_id, image_path, linked_statement_entry_id, created_at
            ) values (?, 0, ?, ?, ?, 'IN', ?, 'commission', 0, null, null, ?, ?)
            """,
            (
                account["id"],
                statement_entry["entry_date"],
                description,
                statement_entry["currency"],
                float(statement_entry["amount"]),
                statement_entry_id,
                utc_timestamp(),
            ),
        )
        expense_entry_id = cursor.lastrowid
    else:
        expense_entry_id = existing_expense["id"]
        affected_accounts.add(existing_expense["account_id"])
        db.execute(
            """
            update expense_entries
            set account_id = ?, entry_date = ?, description = ?, currency = ?, direction = 'IN',
                amount = ?, category = 'commission', linked_statement_entry_id = ?
            where id = ?
            """,
            (
                account["id"],
                statement_entry["entry_date"],
                description,
                statement_entry["currency"],
                float(statement_entry["amount"]),
                statement_entry_id,
                expense_entry_id,
            ),
        )

    affected_accounts.add(account["id"])
    db.execute(
        """
        update statement_entries
        set profit_expense_entry_id = ?, profit_expense_account_id = ?
        where id = ?
        """,
        (expense_entry_id, account["id"], statement_entry_id),
    )
    return affected_accounts


def linked_expense_entry_message(
    db: sqlite3.Connection,
    expense_entry: sqlite3.Row,
) -> str:
    statement_id = expense_entry["linked_statement_entry_id"] if "linked_statement_entry_id" in expense_entry.keys() else None
    if not statement_id:
        return "This expense entry is linked from a client statement row. Edit it from the statement page."
    row = db.execute(
        """
        select se.id, se.source_no, c.name as client_name
        from statement_entries se
        left join clients c on c.id = se.client_id
        where se.id = ?
        """,
        (statement_id,),
    ).fetchone()
    if row is None:
        return "This expense entry is linked from a client statement row. Edit it from the statement page."
    client_label = row["client_name"] or "Client"
    return f"This expense entry is linked from {client_label} row #{row['source_no']}. Edit it from the client statement page."


def annotate_expense_rows(
    db: sqlite3.Connection,
    rows: list[dict],
) -> list[dict]:
    statement_ids = {
        int(row["linked_statement_entry_id"])
        for row in rows
        if row.get("linked_statement_entry_id")
    }
    if not statement_ids:
        return rows
    placeholders = ",".join("?" for _ in statement_ids)
    linked_rows = db.execute(
        f"""
        select se.id as statement_id, se.source_no, se.client_id, c.name as client_name
        from statement_entries se
        left join clients c on c.id = se.client_id
        where se.id in ({placeholders})
        """,
        tuple(statement_ids),
    ).fetchall()
    info_by_statement = {
        int(row["statement_id"]): {
            "source_no": row["source_no"],
            "client_id": row["client_id"],
            "client_name": row["client_name"],
        }
        for row in linked_rows
    }
    for row in rows:
        linked_statement_id = row.get("linked_statement_entry_id")
        if not linked_statement_id:
            row["linked_statement_source_no"] = None
            row["linked_statement_client_id"] = None
            row["linked_statement_client_name"] = None
            continue
        info = info_by_statement.get(int(linked_statement_id), {})
        row["linked_statement_source_no"] = info.get("source_no")
        row["linked_statement_client_id"] = info.get("client_id")
        row["linked_statement_client_name"] = info.get("client_name")
    return rows


def _get_live_usd_cny_rate() -> float | None:
    """Fetch live USD→CNY rate using the existing FX cache/API infrastructure."""
    import time

    now = time.time()
    cached = _fx_cache.get("USD")
    if cached and (now - cached["fetched_at"]) < _FX_CACHE_TTL:
        rate = cached["rates"].get("CNY")
        if rate is not None:
            return float(rate)

    proxies = {"http": FX_PROXY_URL, "https": FX_PROXY_URL} if FX_PROXY_URL else None
    ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")

    try:
        usd_rates = _fetch_moneyconvert(proxies)
        if usd_rates:
            _fx_cache["USD"] = {"rates": usd_rates, "fetched_at": now, "timestamp": ts, "source": "moneyconvert"}
            rate = usd_rates.get("CNY")
            if rate is not None:
                return float(rate)
    except Exception:
        pass

    try:
        rates = _fetch_er_api("USD", proxies)
        if rates:
            _fx_cache["USD"] = {"rates": rates, "fetched_at": now, "timestamp": ts, "source": "exchangerate-api"}
            rate = rates.get("CNY")
            if rate is not None:
                return float(rate)
    except Exception:
        pass

    return None


def exchange_rate_summary() -> dict:
    db = get_db()
    rows = db.execute(
        """
        select transfer_group, min(entry_date) as entry_date, max(exchange_rate) as exchange_rate
        from statement_entries
        where transfer_group is not null
          and currency in ('USD', 'CNY')
          and exchange_rate is not null
        group by transfer_group
        order by entry_date desc, transfer_group desc
        """
    ).fetchall()
    rates = [float(row["exchange_rate"]) for row in rows if row["exchange_rate"] is not None]
    latest = rows[0] if rows else None
    average_rate = (sum(rates) / len(rates)) if rates else None
    fx_source = get_setting("fx_rate_source", "live")
    stored_live = get_setting("fx_live_rate", "")
    live_rate = float(stored_live) if stored_live else None
    live_rate_date = get_setting("fx_live_rate_updated_at", "")
    display_rate = (live_rate if fx_source == "live" and live_rate else None) or average_rate
    return {
        "count": len(rows),
        "latest_rate": float(latest["exchange_rate"]) if latest and latest["exchange_rate"] is not None else None,
        "latest_date": latest["entry_date"] if latest else "",
        "average_rate": average_rate,
        "live_rate": live_rate,
        "live_rate_date": live_rate_date,
        "display_rate": display_rate,
        "fx_source": fx_source,
    }


def company_status(stats: dict, bank_tots: dict, sup_tots: dict) -> dict:
    # Company balance = bank cash - net client balances - supplier debt.
    # Client balances are stored as IN - OUT, so a negative balance means
    # clients owe you money and should increase company balance.
    bank_usd = bank_tots["total_usd"]
    bank_cny = bank_tots["total_cny"]
    stmt_usd = stats["total_usd_balance"]
    stmt_cny = stats["total_cny_balance"]
    sup_usd = sup_tots["total_usd"]
    sup_cny = sup_tots["total_cny"]
    receivable_usd = max(-stmt_usd, 0)
    receivable_cny = max(-stmt_cny, 0)
    balance_usd = bank_usd - stmt_usd - sup_usd
    balance_cny = bank_cny - stmt_cny - sup_cny
    usd_state = "positive" if balance_usd > 0.005 else "negative" if balance_usd < -0.005 else "zero"
    cny_state = "positive" if balance_cny > 0.005 else "negative" if balance_cny < -0.005 else "zero"
    states = {usd_state, cny_state}
    if states == {"positive"}:
        overall_state = "positive"
        overall_label = "Positive Balance"
    elif states == {"negative"}:
        overall_state = "negative"
        overall_label = "Negative Balance"
    elif states == {"zero"}:
        overall_state = "zero"
        overall_label = "Balanced"
    else:
        overall_state = "mixed"
        overall_label = "Mixed Balance"
    return {
        "bank_usd": bank_usd,
        "bank_cny": bank_cny,
        "stmt_usd": stmt_usd,
        "stmt_cny": stmt_cny,
        "sup_usd": sup_usd,
        "sup_cny": sup_cny,
        "receivable_usd": receivable_usd,
        "receivable_cny": receivable_cny,
        "balance_usd": balance_usd,
        "balance_cny": balance_cny,
        "overall_state": overall_state,
        "overall_label": overall_label,
    }


def render_index(error: str | None = None) -> str:
    balances = bank_balance_list()
    suppliers = supplier_balance_list()
    stats = dashboard_stats()
    bank_tots = bank_balance_totals(balances)
    sup_tots = supplier_balance_totals(suppliers)
    fx_summary = exchange_rate_summary()
    return render_template(
        "index.html",
        clients=client_list(),
        groups=grouped_client_list(),
        stats=stats,
        bank_balances=balances,
        bank_totals=bank_tots,
        suppliers=suppliers,
        supplier_totals=sup_tots,
        status=company_status(stats, bank_tots, sup_tots),
        fx_summary=fx_summary,
        expense_accounts=expense_dashboard_data(),
        error=error,
    )


def import_client_csv_rows(client_id: int, rows: list[dict[str, str]]) -> None:
    db = get_db()
    seen_transfer_groups: set[str] = set()
    for row in rows:
        entry_date = (row.get("date") or row.get("entry_date") or "").strip()
        description = (row.get("description") or "").strip()
        currency = (row.get("currency") or "").strip().upper()
        direction = (row.get("direction") or "").strip().upper()
        amount_raw = (row.get("amount") or "").strip()
        kind = (row.get("kind") or "movement").strip().lower() or "movement"
        category_hint = (row.get("category_hint") or UNCATEGORIZED).strip() or UNCATEGORIZED
        transfer_group = (row.get("transfer_group") or "").strip() or None
        exchange_rate_raw = (row.get("exchange_rate") or "").strip()
        source_no_raw = (row.get("source_no") or "").strip()

        if not entry_date or not description or currency not in {"USD", "CNY"} or direction not in {"IN", "OUT"}:
            raise ValueError("CSV rows must include valid date, description, currency, and direction values.")
        amount = float(amount_raw)
        source_no = int(float(source_no_raw)) if source_no_raw else 0
        exchange_rate = float(exchange_rate_raw) if exchange_rate_raw else None
        if kind not in {"movement", "transfer"}:
            kind = "movement"

        db.execute(
            """
            insert into statement_entries (
                client_id, source_no, entry_date, description, currency, direction,
                amount, kind, category_hint, transfer_group, exchange_rate
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_id,
                source_no,
                entry_date,
                description,
                currency,
                direction,
                amount,
                kind,
                category_hint,
                transfer_group,
                exchange_rate,
            ),
        )
        if transfer_group:
            seen_transfer_groups.add(transfer_group)
    for transfer_group in seen_transfer_groups:
        sync_exchange_group(db, transfer_group)
    db.commit()
    resequence_client_entries(client_id)


ENTRY_ORDER = "entry_date, id"


def resequence_client_entries(client_id: int) -> None:
    db = get_db()
    entry_ids = db.execute(
        f"select id from statement_entries where client_id = ? order by {ENTRY_ORDER}",
        (client_id,),
    ).fetchall()
    for index, row in enumerate(entry_ids, start=1):
        db.execute("update statement_entries set source_no = ? where id = ?", (index, row["id"]))
    db.commit()


def resequence_all_clients() -> None:
    """Resequence source_no for every client so entries are always sorted by date."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    clients = db.execute("select id from clients").fetchall()
    for client in clients:
        cid = client["id"]
        entry_ids = db.execute(
            f"select id from statement_entries where client_id = ? order by {ENTRY_ORDER}",
            (cid,),
        ).fetchall()
        for index, row in enumerate(entry_ids, start=1):
            db.execute("update statement_entries set source_no = ? where id = ?", (index, row["id"]))
    db.commit()
    db.close()


def running_balances(entries: list[sqlite3.Row]) -> list[dict]:
    usd = 0.0
    cny = 0.0
    rows: list[dict] = []
    for entry in entries:
        amount = float(entry["amount"])
        if entry["currency"] == "USD":
            usd += amount if entry["direction"] == "IN" else -amount
        if entry["currency"] == "CNY":
            cny += amount if entry["direction"] == "IN" else -amount
        rows.append(
            {
                **dict(entry),
                "running_usd": usd,
                "running_cny": cny,
                "exchange_rate": entry["exchange_rate"] if "exchange_rate" in entry.keys() else None,
            }
        )
    transfer_groups: dict[str, dict[str, float]] = {}
    for row in rows:
        transfer_group = row.get("transfer_group")
        if not transfer_group:
            continue
        group = transfer_groups.setdefault(transfer_group, {"usd": 0.0, "cny": 0.0})
        if row["currency"] == "USD":
            group["usd"] += abs(float(row["amount"]))
        elif row["currency"] == "CNY":
            group["cny"] += abs(float(row["amount"]))
    for row in rows:
        transfer_group = row.get("transfer_group")
        if not transfer_group:
            continue
        group = transfer_groups.get(transfer_group, {})
        usd_amount = group.get("usd", 0.0)
        cny_amount = group.get("cny", 0.0)
        if usd_amount > 0 and cny_amount > 0 and not row.get("exchange_rate"):
            row["exchange_rate"] = cny_amount / usd_amount
    return rows


def row_to_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "client_id": row["client_id"],
        "source_no": row["source_no"],
        "entry_date": row["entry_date"],
        "description": row["description"],
        "currency": row["currency"],
        "direction": row["direction"],
        "amount": row["amount"],
        "kind": row["kind"],
        "category_hint": row["category_hint"],
        "transfer_group": row["transfer_group"],
        "exchange_rate": row["exchange_rate"] if "exchange_rate" in row.keys() else None,
        "linked_entry_id": row["linked_entry_id"] if "linked_entry_id" in row.keys() else None,
        "commission_source_entry_id": row["commission_source_entry_id"] if "commission_source_entry_id" in row.keys() else None,
        "profit_expense_entry_id": row["profit_expense_entry_id"] if "profit_expense_entry_id" in row.keys() else None,
        "profit_expense_account_id": row["profit_expense_account_id"] if "profit_expense_account_id" in row.keys() else None,
        "image_path": row["image_path"] if "image_path" in row.keys() else None,
    }


def make_transfer_group() -> str:
    return f"FX-{utc_now().strftime('%Y%m%d%H%M%S')}"


def save_upload_image(file_storage) -> str | None:
    """Save an uploaded image file and return its filename, or None."""
    if not file_storage or not file_storage.filename:
        return None
    ext = Path(file_storage.filename).suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        return None
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex}{ext}"
    file_storage.save(UPLOAD_DIR / filename)
    return filename


def record_event(client_id: int, entry_id: int | None, action: str, payload: dict) -> None:
    db = get_db()
    db.execute(
        """
        insert into statement_entry_events (client_id, entry_id, action, payload, created_at)
        values (?, ?, ?, ?, ?)
        """,
        (
            client_id,
            entry_id,
            action,
            json.dumps(payload),
            utc_timestamp(),
        ),
    )
    db.commit()


def latest_undo_event(client_id: int) -> sqlite3.Row | None:
    db = get_db()
    return db.execute(
        """
        select *
        from statement_entry_events
        where client_id = ? and undone_at is null
        order by id desc
        limit 1
        """,
        (client_id,),
    ).fetchone()


def recent_events(client_id: int, limit: int = 8) -> list[dict]:
    db = get_db()
    rows = db.execute(
        """
        select *
        from statement_entry_events
        where client_id = ?
        order by id desc
        limit ?
        """,
        (client_id, limit),
    ).fetchall()
    events: list[dict] = []
    for row in rows:
        payload = json.loads(row["payload"])
        description = ""
        if row["action"] == "add":
            description = payload.get("entry", {}).get("description", "")
        elif row["action"] == "delete":
            description = payload.get("entry", {}).get("description", "")
        elif row["action"] == "edit":
            description = payload.get("after", {}).get("description") or payload.get("before", {}).get("description", "")
        elif row["action"] == "commission":
            description = payload.get("entry", {}).get("description", "Commission")
        elif row["action"] == "exchange":
            description = payload.get("source_description", "Currency exchange")
        events.append(
            {
                "id": row["id"],
                "action": row["action"],
                "description": description,
                "created_at": normalize_utc_timestamp(row["created_at"]),
                "undone": bool(row["undone_at"]),
            }
        )
    return events


TYPE_OPTIONS = [
    ("movement", "Regular"),
    ("transfer", "FX / Transfer"),
]


CATEGORY_OPTIONS = [
    (UNCATEGORIZED, "Uncategorized"),
    ("client_receipt", "Client Receipt"),
    ("factory_payment", "Factory Payment"),
    ("commission", "Commission"),
    ("shipping_expense", "Shipping"),
    ("travel_expense", "Travel"),
    ("internal_transfer", "Internal Transfer"),
    ("fx_transfer", "Currency Exchange"),
]


ALL_EXPENSE_CURRENCIES = ["USD", "CNY", "EGP"]

EXPENSE_CATEGORIES = [
    ("general", "General"),
    ("rent", "Rent"),
    ("salary", "Salary"),
    ("insurance", "Insurance"),
    ("utilities", "Utilities"),
    ("cleaning", "Cleaning"),
    ("parking", "Parking"),
    ("commission", "Commission"),
    ("deposit", "Deposit"),
    ("other", "Other"),
]


# --- Expense helper functions ---

def expense_running_balances(entries, enabled_currencies):
    """Multi-currency running balance for expense entries."""
    balances = {c: 0.0 for c in enabled_currencies}
    rows = []
    for entry in entries:
        amount = float(entry["amount"])
        cur = entry["currency"]
        if cur in balances:
            if entry["direction"] == "IN":
                balances[cur] += amount
            else:
                balances[cur] -= amount
        row = dict(entry)
        row["running_balances"] = dict(balances)
        rows.append(row)
    return rows


def resequence_expense_entries(account_id):
    db = get_db()
    entry_ids = db.execute(
        "select id from expense_entries where account_id = ? order by entry_date, id",
        (account_id,),
    ).fetchall()
    for index, row in enumerate(entry_ids, start=1):
        db.execute("update expense_entries set seq_no = ? where id = ?", (index, row["id"]))
    db.commit()


def expense_record_event(account_id, entry_id, action, payload):
    db = get_db()
    db.execute(
        "insert into expense_events (account_id, entry_id, action, payload, created_at) values (?, ?, ?, ?, ?)",
        (account_id, entry_id, action, json.dumps(payload), utc_timestamp()),
    )
    db.commit()


def expense_entry_to_dict(row):
    return {
        "id": row["id"],
        "account_id": row["account_id"],
        "seq_no": row["seq_no"],
        "entry_date": row["entry_date"],
        "description": row["description"],
        "currency": row["currency"],
        "direction": row["direction"],
        "amount": row["amount"],
        "category": row["category"],
        "is_recurring": row["is_recurring"],
        "template_id": row["template_id"],
        "image_path": row["image_path"],
        "linked_statement_entry_id": row["linked_statement_entry_id"] if "linked_statement_entry_id" in row.keys() else None,
    }


def generate_recurring_expenses(account_id):
    """Auto-generate entries from active recurring templates up to today."""
    db = get_db()
    today = china_today()
    current_ym = today.strftime("%Y-%m")

    templates = db.execute(
        "select * from recurring_expense_templates where account_id = ? and is_active = 1",
        (account_id,),
    ).fetchall()

    generated_count = 0
    for tpl in templates:
        every_n = tpl["every_n_months"] if "every_n_months" in tpl.keys() else 1
        last_gen = tpl["last_generated"] or ""
        created_at = tpl["created_at"]
        if created_at:
            try:
                start_ym = created_at[:7]
            except Exception:
                start_ym = current_ym
        else:
            start_ym = current_ym

        if last_gen:
            year, month = map(int, last_gen.split("-"))
            month += every_n
            while month > 12:
                month -= 12
                year += 1
            start_ym = f"{year:04d}-{month:02d}"

        year, month = map(int, start_ym.split("-"))
        end_year, end_month = map(int, current_ym.split("-"))

        while (year, month) <= (end_year, end_month):
            ym = f"{year:04d}-{month:02d}"
            day = min(tpl["day_of_month"], 28)
            entry_date = f"{ym}-{day:02d}"

            try:
                entry_d = date.fromisoformat(entry_date)
                if entry_d > today:
                    month += every_n
                    while month > 12:
                        month -= 12
                        year += 1
                    continue
            except Exception:
                pass

            existing = db.execute(
                "select id from expense_entries where account_id = ? and template_id = ? and entry_date = ?",
                (account_id, tpl["id"], entry_date),
            ).fetchone()

            if existing is None:
                deleted_check = db.execute(
                    """select id from expense_events
                    where account_id = ? and action = 'delete'
                    and json_extract(payload, '$.entry.template_id') = ?
                    and json_extract(payload, '$.entry.entry_date') = ?""",
                    (account_id, tpl["id"], entry_date),
                ).fetchone()

                if deleted_check is None:
                    db.execute(
                        """insert into expense_entries
                        (account_id, seq_no, entry_date, description, currency, direction,
                         amount, category, is_recurring, template_id, created_at)
                        values (?, 0, ?, ?, ?, ?, ?, ?, 1, ?, ?)""",
                        (account_id, entry_date, tpl["description"], tpl["currency"],
                         tpl["direction"], tpl["amount"], tpl["category"], tpl["id"], utc_timestamp()),
                    )
                    generated_count += 1

            db.execute(
                "update recurring_expense_templates set last_generated = ? where id = ?",
                (ym, tpl["id"]),
            )

            month += every_n
            while month > 12:
                month -= 12
                year += 1

    db.commit()
    if generated_count > 0:
        resequence_expense_entries(account_id)
    return generated_count


def expense_account_summary(db, account_id):
    """Get summary totals for an expense account."""
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if not account:
        return None
    currencies = [c.strip() for c in account["enabled_currencies"].split(",") if c.strip()]
    entry_count = db.execute("select count(*) from expense_entries where account_id = ?", (account_id,)).fetchone()[0]
    totals = {}
    for cur in currencies:
        total_in = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='IN'",
            (account_id, cur),
        ).fetchone()[0]
        total_out = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='OUT'",
            (account_id, cur),
        ).fetchone()[0]
        totals[cur] = {"in": total_in, "out": total_out, "balance": total_in - total_out}
    return {
        "id": account["id"],
        "name": account["name"],
        "currencies": currencies,
        "entry_count": entry_count,
        "totals": totals,
        "created_at": account["created_at"],
    }


def expense_dashboard_data():
    db = get_db()
    accounts = db.execute("select * from expense_accounts order by name").fetchall()
    result = []
    for acc in accounts:
        summary = expense_account_summary(db, acc["id"])
        if summary:
            result.append(summary)
    return result


# --- Auth routes ---

@app.route("/login", methods=["GET"])
def login_page():
    if "user_id" in session:
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login_action():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    db = get_db()
    user = db.execute("select * from users where username = ?", (username,)).fetchone()
    if user and user["is_active"] and check_password_hash(user["password_hash"], password):
        session.clear()
        session.permanent = True
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user["role"]
        get_csrf_token()
        db.execute("update users set last_login = ? where id = ?",
                   (datetime.now(timezone.utc).isoformat(), user["id"]))
        db.commit()
        if user["must_change_password"]:
            return redirect(url_for("change_password"))
        return redirect(url_for("index"))
    flash("Invalid username or password", "error")
    return render_template("login.html")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login_page"))


@app.route("/change-password", methods=["GET", "POST"])
def change_password():
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not g.user["must_change_password"] and not check_password_hash(g.user["password_hash"], current):
            flash("Current password is incorrect", "error")
        elif len(new_pw) < 6:
            flash("Password must be at least 6 characters", "error")
        elif new_pw != confirm:
            flash("Passwords do not match", "error")
        else:
            db = get_db()
            db.execute("update users set password_hash = ?, must_change_password = 0 where id = ?",
                       (generate_password_hash(new_pw), g.user["id"]))
            db.commit()
            flash("Password changed successfully", "success")
            return redirect(url_for("index"))
    return render_template("change_password.html", forced=g.user["must_change_password"])


@app.route("/reset-secret-key", methods=["POST"])
@admin_required
def reset_secret_key():
    """Rotate the session secret and invalidate existing sessions."""
    if not RESET_SECRET_TOKEN:
        abort(404)
    token = request.form.get("reset_secret_token", "").strip()
    if token != RESET_SECRET_TOKEN:
        abort(403)
    db = get_db()
    new_key = secrets.token_hex(32)
    db.execute("insert or replace into app_settings(key, value) values ('secret_key', ?)", (new_key,))
    db.commit()
    refresh_app_secret_key(db)
    session.clear()
    flash("Session secret rotated. Sign in again.", "success")
    return redirect(url_for("login_page"))


# --- Admin routes ---

@app.route("/admin/users")
@admin_required
def admin_users():
    db = get_db()
    users = db.execute("select * from users order by id").fetchall()
    return render_template("admin.html", users=users)


@app.route("/admin/users/add", methods=["POST"])
@admin_required
def admin_add_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "user")
    if not username or not password:
        flash("Username and password are required", "error")
        return redirect(url_for("admin_users"))
    if len(password) < 6:
        flash("Password must be at least 6 characters", "error")
        return redirect(url_for("admin_users"))
    if role not in ("admin", "user"):
        role = "user"
    db = get_db()
    existing = db.execute("select id from users where username = ?", (username,)).fetchone()
    if existing:
        flash(f"Username '{username}' already exists", "error")
        return redirect(url_for("admin_users"))
    db.execute(
        "insert into users(username, password_hash, role, must_change_password, created_at) values (?, ?, ?, 1, ?)",
        (username, generate_password_hash(password), role, datetime.now(timezone.utc).isoformat()),
    )
    db.commit()
    flash(f"User '{username}' created. They must change password on first login.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@admin_required
def admin_toggle_user(user_id):
    if user_id == g.user["id"]:
        flash("Cannot deactivate yourself", "error")
        return redirect(url_for("admin_users"))
    db = get_db()
    user = db.execute("select * from users where id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    db.execute("update users set is_active = ? where id = ?", (0 if user["is_active"] else 1, user_id))
    db.commit()
    flash(f"User '{user['username']}' {'activated' if not user['is_active'] else 'deactivated'}", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/role", methods=["POST"])
@admin_required
def admin_change_role(user_id):
    if user_id == g.user["id"]:
        flash("Cannot change your own role", "error")
        return redirect(url_for("admin_users"))
    db = get_db()
    user = db.execute("select * from users where id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    new_role = "admin" if user["role"] == "user" else "user"
    db.execute("update users set role = ? where id = ?", (new_role, user_id))
    db.commit()
    flash(f"User '{user['username']}' is now {new_role}", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/reset-password", methods=["POST"])
@admin_required
def admin_reset_password(user_id):
    new_pw = request.form.get("new_password", "").strip()
    if not new_pw or len(new_pw) < 6:
        flash("Password must be at least 6 characters", "error")
        return redirect(url_for("admin_users"))
    db = get_db()
    user = db.execute("select * from users where id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    db.execute("update users set password_hash = ?, must_change_password = 1 where id = ?",
               (generate_password_hash(new_pw), user_id))
    db.commit()
    flash(f"Password reset for '{user['username']}'. They must change it on next login.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def admin_delete_user(user_id):
    if user_id == g.user["id"]:
        flash("Cannot delete yourself", "error")
        return redirect(url_for("admin_users"))
    db = get_db()
    user = db.execute("select * from users where id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    db.execute("delete from users where id = ?", (user_id,))
    db.commit()
    flash(f"User '{user['username']}' deleted", "success")
    return redirect(url_for("admin_users"))


# --- Main routes ---

@app.route("/")
def index():
    return render_index()


@app.route("/clients/new", methods=["POST"])
def new_client():
    name = request.form.get("name", "").strip()
    if not name:
        abort(400)
    db = get_db()
    db.execute("insert into clients(name) values (?)", (name,))
    db.commit()
    client_id = db.execute("select id from clients where name = ?", (name,)).fetchone()["id"]
    return redirect(url_for("client_statement", client_id=client_id))


@app.route("/clients/import", methods=["POST"])
def import_client_csv():
    name = request.form.get("name", "").strip()
    upload = request.files.get("statement_csv")
    if not name:
        return render_index("Client name is required for CSV import."), 400
    if upload is None or not upload.filename:
        return render_index("Choose a CSV file to import."), 400

    try:
        content = upload.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        return render_index("CSV must be UTF-8 encoded."), 400

    reader = csv.DictReader(io.StringIO(content))
    rows = [dict(row) for row in reader]
    if not rows:
        return render_index("CSV file is empty."), 400

    required_any = {"date", "entry_date"}
    fieldnames = set(reader.fieldnames or [])
    if not (required_any & fieldnames) or "description" not in fieldnames or "currency" not in fieldnames or "direction" not in fieldnames or "amount" not in fieldnames:
        return render_index("CSV must include date, description, currency, direction, and amount columns."), 400

    db = get_db()
    existing = db.execute("select id from clients where name = ?", (name,)).fetchone()
    if existing is not None:
        return render_index(f'Client "{name}" already exists.'), 400

    try:
        db.execute("insert into clients(name) values (?)", (name,))
        db.commit()
        client_id = db.execute("select id from clients where name = ?", (name,)).fetchone()["id"]
        import_client_csv_rows(client_id, rows)
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        client = db.execute("select id from clients where name = ?", (name,)).fetchone()
        if client is not None:
            db.execute("delete from clients where id = ?", (client["id"],))
            db.commit()
        return render_index(f"CSV import failed: {exc}"), 400

    return redirect(url_for("client_statement", client_id=client_id))


@app.route("/clients/<int:client_id>")
def client_statement(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    commission_expense_accounts = expense_account_options(db)
    default_profit_account_id = default_profit_expense_account_id(commission_expense_accounts)

    filters = {
        "q": request.args.get("q", "").strip(),
        "currency": request.args.get("currency", "").strip(),
        "category": request.args.get("category", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
    }
    view_mode = request.args.get("view", "combined").strip() or "combined"
    if view_mode not in {"combined", "split"}:
        view_mode = "combined"
    page_arg = request.args.get("page", type=int)
    page = max(page_arg or 1, 1)
    per_page_raw = request.args.get("per_page", "10").strip() or "10"
    per_page = "all" if per_page_raw == "all" else max(int(per_page_raw), 1)
    clauses = ["client_id = ?"]
    params: list[object] = [client_id]
    if filters["q"]:
        clauses.append("(description like ? or cast(amount as text) like ? or cast(source_no as text) like ? or entry_date like ? or currency like ? or direction like ? or kind like ? or coalesce(category_hint,'') like ? or coalesce(transfer_group,'') like ?)")
        q_pat = f'%{filters["q"]}%'
        params.extend([q_pat] * 9)
    if filters["currency"]:
        clauses.append("currency = ?")
        params.append(filters["currency"])
    if filters["category"]:
        clauses.append("category_hint = ?")
        params.append(filters["category"])
    if filters["date_from"]:
        clauses.append("entry_date >= ?")
        params.append(filters["date_from"])
    if filters["date_to"]:
        clauses.append("entry_date <= ?")
        params.append(filters["date_to"])

    where_sql = " and ".join(clauses)
    query = f"""
        select *
        from statement_entries
        where {where_sql}
        order by entry_date, id
    """
    all_entries = db.execute(query, params).fetchall()
    total_rows = len(all_entries)
    if per_page == "all":
        total_pages = 1
        page = 1
        entries = all_entries
    else:
        total_pages = max((total_rows + per_page - 1) // per_page, 1)
        if page > total_pages:
            page = total_pages
        pages: list[list[sqlite3.Row]] = []
        remainder = total_rows % per_page
        first_chunk = remainder if remainder else per_page
        start_index = 0
        if total_rows > 0:
            pages.append(all_entries[:first_chunk])
            start_index = first_chunk
        while start_index < total_rows:
            pages.append(all_entries[start_index:start_index + per_page])
            start_index += per_page
        if pages:
            pages = list(reversed(pages))
        entries = pages[page - 1] if pages else []

    rows = statement_rows_with_commission_state(db, client_id, all_entries)
    current_usd_balance = rows[-1]["running_usd"] if rows else 0.0
    current_cny_balance = rows[-1]["running_cny"] if rows else 0.0
    fx_summary = exchange_rate_summary()
    approx_total_rmb = None
    if fx_summary.get("display_rate") is not None:
        approx_total_rmb = current_usd_balance * fx_summary["display_rate"] + current_cny_balance
    if per_page != "all":
        pages: list[list[dict]] = []
        remainder = total_rows % per_page
        first_chunk = remainder if remainder else per_page
        start_index = 0
        if total_rows > 0:
            pages.append(rows[:first_chunk])
            start_index = first_chunk
        while start_index < total_rows:
            pages.append(rows[start_index:start_index + per_page])
            start_index += per_page
        if pages:
            pages = list(reversed(pages))
        rows = pages[page - 1] if pages else []
    return render_template(
        "client.html",
        client=client,
        rows=rows,
        current_usd_balance=current_usd_balance,
        current_cny_balance=current_cny_balance,
        approx_total_rmb=approx_total_rmb,
        fx_summary=fx_summary,
        latest_event=latest_undo_event(client_id),
        recent_events=recent_events(client_id),
        filters=filters,
        view_mode=view_mode,
        page=page,
        per_page=per_page,
        per_page_raw=per_page_raw,
        total_rows=total_rows,
        total_pages=total_pages,
        default_date=china_today().isoformat(),
        type_options=TYPE_OPTIONS,
        category_options=CATEGORY_OPTIONS,
        commission_expense_accounts=commission_expense_accounts,
        default_profit_account_id=default_profit_account_id,
        quick_submit_id=request.args.get("quick_submit_id", type=int),
        quick_submit=_get_quick_submit(request.args.get("quick_submit_id", type=int)),
    )


def _get_quick_submit(qs_id: int | None) -> dict | None:
    if not qs_id:
        return None
    row = get_db().execute("select * from quick_submits where id = ?", (qs_id,)).fetchone()
    if not row:
        return None
    return {"image_path": row["image_path"], "description": row["description"], "amount": row["amount"]}


@app.route("/clients/<int:client_id>/entries", methods=["POST"])
def add_entry(client_id: int):
    db = get_db()
    client = db.execute("select id from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    entry_date = request.form["entry_date"]
    description = request.form["description"].strip()
    currency = request.form["currency"]
    direction = request.form["direction"]
    amount = float(request.form["amount"])
    kind = request.form["kind"]
    category_hint = request.form["category_hint"]
    transfer_group = request.form.get("transfer_group", "").strip() or None
    add_to_company_profit = request.form.get("add_to_company_profit") == "1"
    profit_expense_account_id = request.form.get("profit_expense_account_id", type=int)
    try:
        validate_profit_expense_selection(
            db,
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
            category_hint=category_hint,
            currency=currency,
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(_return_url(client_id, "entry-new"))
    image_filename = save_upload_image(request.files.get("image"))
    # If no new file but an existing image from quick submit
    if not image_filename:
        image_filename = request.form.get("existing_image_path") or None
    db.execute(
        """
        insert into statement_entries (
            client_id, source_no, entry_date, description, currency, direction,
            amount, kind, category_hint, transfer_group, image_path
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            client_id,
            0,
            entry_date,
            description,
            currency,
            direction,
            amount,
            kind,
            category_hint,
            transfer_group,
            image_filename,
        ),
    )
    entry = db.execute(
        "select * from statement_entries where client_id = ? order by id desc limit 1",
        (client_id,),
    ).fetchone()
    affected_expense_accounts: set[int] = set()
    if entry is not None:
        affected_expense_accounts = sync_statement_profit_entry(
            db,
            entry["id"],
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
        )
    sync_exchange_group(db, transfer_group)
    db.commit()
    entry = db.execute(
        "select * from statement_entries where client_id = ? order by id desc limit 1",
        (client_id,),
    ).fetchone()
    if entry is not None:
        record_event(client_id, entry["id"], "add", {"entry": row_to_dict(entry)})
    # Mark quick submit as processed if applicable
    qs_id = request.form.get("quick_submit_id", type=int)
    if qs_id and entry is not None:
        db.execute(
            "update quick_submits set status='processed', processed_at=?, created_entry_id=? where id=?",
            (utc_timestamp(), entry["id"], qs_id),
        )
    db.commit()
    resequence_client_entries(client_id)
    for account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(account_id)
    return redirect(_return_url(client_id, "entry-new"))


@app.route("/clients/<int:client_id>/exchange", methods=["POST"])
def exchange_balance(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)

    exchange_date = request.form["exchange_date"]
    usd_amount_raw = request.form.get("usd_amount", "").strip()
    usd_amount = float(usd_amount_raw) if usd_amount_raw else 0.0
    cny_amount_raw = request.form.get("cny_amount", "").strip()
    cny_amount = float(cny_amount_raw) if cny_amount_raw else 0.0
    submitted_rate_raw = request.form.get("exchange_rate", "").strip()
    submitted_rate = float(submitted_rate_raw) if submitted_rate_raw else None
    fee_enabled = request.form.get("apply_fee") == "1"
    fee_rate = 0.004 if fee_enabled else 0.0
    note = request.form.get("exchange_note", "").strip()
    transfer_group = make_transfer_group()
    usd_description = note or "USD to RMB exchange"
    cny_description = note or "USD to RMB exchange"
    if submitted_rate and cny_amount > 0 and usd_amount <= 0:
        divisor = submitted_rate * (1 - fee_rate)
        if divisor > 0:
            usd_amount = round(cny_amount / divisor, 2)
    elif submitted_rate and usd_amount > 0 and cny_amount <= 0:
        # Calculate CNY from rate, then apply fee if enabled
        cny_amount = round(usd_amount * submitted_rate, 2)
        if fee_enabled:
            cny_amount = round(cny_amount * (1 - fee_rate), 2)

    balance_rows = running_balances(
        db.execute(
            "select * from statement_entries where client_id = ? order by entry_date, id",
            (client_id,),
        ).fetchall()
    )
    available_usd = balance_rows[-1]["running_usd"] if balance_rows else 0.0
    if usd_amount <= 0:
        flash(f"USD amount must be greater than zero.", "error")
        return redirect(url_for("client_statement", client_id=client_id))
    if round(usd_amount, 2) > round(available_usd, 2):
        flash(f"USD amount (${usd_amount:,.2f}) exceeds available balance (${available_usd:,.2f}).", "error")
        return redirect(url_for("client_statement", client_id=client_id))
    if cny_amount <= 0:
        flash("RMB amount could not be calculated. Please enter an exchange rate or RMB amount.", "error")
        return redirect(url_for("client_statement", client_id=client_id))
    effective_rate = cny_amount / usd_amount

    db.execute(
        """
        insert into statement_entries (
            client_id, source_no, entry_date, description, currency, direction,
            amount, kind, category_hint, transfer_group, exchange_rate
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            client_id,
            0,
            exchange_date,
            usd_description,
            "USD",
            "OUT",
            usd_amount,
            "transfer",
            "fx_transfer",
            transfer_group,
            effective_rate,
        ),
    )
    usd_entry = db.execute("select * from statement_entries where client_id = ? order by id desc limit 1", (client_id,)).fetchone()
    db.execute(
        """
        insert into statement_entries (
            client_id, source_no, entry_date, description, currency, direction,
            amount, kind, category_hint, transfer_group, exchange_rate
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            client_id,
            0,
            exchange_date,
            cny_description,
            "CNY",
            "IN",
            cny_amount,
            "transfer",
            "fx_transfer",
            transfer_group,
            effective_rate,
        ),
    )
    cny_entry = db.execute("select * from statement_entries where client_id = ? order by id desc limit 1", (client_id,)).fetchone()
    sync_exchange_group(db, transfer_group)
    db.commit()
    if usd_entry is not None and cny_entry is not None:
        record_event(
            client_id,
            usd_entry["id"],
            "exchange",
            {
                "source_description": note or "USD to RMB exchange",
                "transfer_group": transfer_group,
                "fee_rate": fee_rate,
                "entries": [row_to_dict(usd_entry), row_to_dict(cny_entry)],
            },
        )
    resequence_client_entries(client_id)
    return redirect(_return_url(client_id, "exchange-balance"))


@app.route("/clients/<int:client_id>/commission", methods=["POST"])
def create_commission(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)

    source_entry_id = request.form.get("source_entry_id", type=int)
    source_entry = db.execute("select * from statement_entries where id = ?", (source_entry_id,)).fetchone()
    source_anchor = f"entry-{source_entry_id}" if source_entry_id else "table-top"
    error = commission_error_message(db, source_entry, client_id)
    if error:
        flash(error, "error")
        return redirect(_return_url(client_id, source_anchor))

    commission_date = request.form.get("commission_date", "").strip()
    if not commission_date:
        flash("Commission date is required.", "error")
        return redirect(_return_url(client_id, source_anchor))

    percentage_raw = request.form.get("percentage", "").strip()
    note = request.form.get("note", "").strip()
    try:
        create_commission_entry(db, source_entry, commission_date, percentage_raw, note)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(_return_url(client_id, source_anchor))

    flash("Commission entry created.", "success")
    return redirect(_return_url(client_id, source_anchor))


def _return_url(client_id: int, anchor: str = "table-top") -> str:
    """Build redirect URL preserving page/per_page from the form submission."""
    page = request.form.get("return_page", "").strip()
    per_page = request.form.get("return_per_page", "").strip()
    params = {}
    if page:
        params["page"] = page
    if per_page:
        params["per_page"] = per_page
    base = url_for("client_statement", client_id=client_id, **params)
    return f"{base}#{anchor}"


@app.route("/entries/<int:entry_id>/save", methods=["POST"])
def save_entry(entry_id: int):
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        abort(404)
    before = row_to_dict(entry)
    new_entry_date = request.form["entry_date"]
    new_description = request.form["description"].strip()
    new_amount = float(request.form["amount"])
    new_currency = request.form["currency"]
    new_direction = request.form["direction"]
    new_kind = request.form["kind"]
    new_category_hint = request.form["category_hint"]
    new_transfer_group = request.form.get("transfer_group", "").strip() or None
    add_to_company_profit = request.form.get("add_to_company_profit") == "1"
    profit_expense_account_id = request.form.get("profit_expense_account_id", type=int)
    if entry["commission_source_entry_id"]:
        source_entry = db.execute(
            "select currency from statement_entries where id = ?",
            (entry["commission_source_entry_id"],),
        ).fetchone()
        new_currency = source_entry["currency"] if source_entry else entry["currency"]
        new_direction = "OUT"
        new_kind = "movement"
        new_category_hint = "commission"
        new_transfer_group = None
    try:
        validate_profit_expense_selection(
            db,
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
            category_hint=new_category_hint,
            currency=new_currency,
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(_return_url(entry["client_id"], f"entry-{entry_id}"))
    old_image = entry["image_path"] if "image_path" in entry.keys() else None
    new_image = save_upload_image(request.files.get("image"))
    image_path = new_image if new_image else old_image
    if request.form.get("remove_image") == "1":
        image_path = None
    # Delete old image file if replaced or removed
    if old_image and old_image != image_path:
        _delete_image_file(old_image)
    db.execute(
        """
        update statement_entries
        set entry_date = ?, description = ?, currency = ?, direction = ?,
            amount = ?, kind = ?, category_hint = ?, transfer_group = ?, image_path = ?
        where id = ?
        """,
        (
            new_entry_date,
            new_description,
            new_currency,
            new_direction,
            new_amount,
            new_kind,
            new_category_hint,
            new_transfer_group,
            image_path,
            entry_id,
        ),
    )
    db.commit()

    sync_exchange_group(db, before.get("transfer_group"))
    sync_exchange_group(db, new_transfer_group)
    affected_expense_accounts = sync_statement_profit_entry(
        db,
        entry_id,
        enabled=add_to_company_profit,
        account_id=profit_expense_account_id,
    )
    db.commit()

    updated = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if updated is not None:
        record_event(entry["client_id"], entry_id, "edit", {"before": before, "after": row_to_dict(updated)})
    resequence_client_entries(entry["client_id"])
    for account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(account_id)

    # Return JSON for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        updated_row = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
        return jsonify({
            "ok": True,
            "entry": {
                "id": entry_id,
                "entry_date": updated_row["entry_date"],
                "source_no": updated_row["source_no"],
                "description": updated_row["description"],
                "currency": updated_row["currency"],
                "direction": updated_row["direction"],
                "amount": updated_row["amount"],
                "kind": updated_row["kind"],
                "category_hint": updated_row["category_hint"] or "uncategorized",
                "transfer_group": updated_row["transfer_group"] or "",
                "commission_source_entry_id": updated_row["commission_source_entry_id"],
                "profit_expense_entry_id": updated_row["profit_expense_entry_id"],
                "profit_expense_account_id": updated_row["profit_expense_account_id"],
                "image_path": updated_row["image_path"] or "",
            }
        })
    return redirect(_return_url(entry["client_id"], f"entry-{entry_id}"))


@app.route("/entries/<int:entry_id>/delete", methods=["POST"])
def delete_entry(entry_id: int):
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        abort(404)
    linked_commission = commission_child_entry(db, entry_id)
    if linked_commission is not None:
        flash("Delete the generated commission entry first.", "error")
        return redirect(_return_url(entry["client_id"], f"entry-{entry_id}"))
    record_event(entry["client_id"], entry_id, "delete", {"entry": row_to_dict(entry)})
    affected_expense_accounts = sync_statement_profit_entry(
        db,
        entry_id,
        enabled=False,
        account_id=None,
    )
    _delete_image_file(entry["image_path"] if "image_path" in entry.keys() else None)
    db.execute("delete from statement_entries where id = ?", (entry_id,))
    db.commit()
    resequence_client_entries(entry["client_id"])
    for account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(account_id)
    return redirect(_return_url(entry["client_id"]))


@app.route("/clients/<int:client_id>/undo", methods=["POST"])
def undo_last_change(client_id: int):
    db = get_db()
    event = latest_undo_event(client_id)
    if event is None:
        return redirect(url_for("client_statement", client_id=client_id))

    payload = json.loads(event["payload"])
    affected_expense_accounts: set[int] = set()
    if event["action"] == "add":
        entry_id = payload.get("entry", {}).get("id")
        current_entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
        if current_entry is not None:
            affected_expense_accounts.update(
                sync_statement_profit_entry(db, current_entry["id"], enabled=False, account_id=None)
            )
            db.execute("delete from statement_entries where id = ?", (current_entry["id"],))
    elif event["action"] == "commission":
        entry_id = payload.get("entry", {}).get("id")
        current_entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
        if current_entry is not None:
            affected_expense_accounts.update(
                sync_statement_profit_entry(db, current_entry["id"], enabled=False, account_id=None)
            )
            db.execute("delete from statement_entries where id = ?", (current_entry["id"],))
    elif event["action"] == "edit":
        before = payload.get("before", {})
        db.execute(
            """
            update statement_entries
            set entry_date = ?, description = ?, currency = ?, direction = ?,
                amount = ?, kind = ?, category_hint = ?, transfer_group = ?, exchange_rate = ?,
                image_path = ?, commission_source_entry_id = ?
            where id = ?
            """,
            (
                before.get("entry_date"),
                before.get("description"),
                before.get("currency"),
                before.get("direction"),
                before.get("amount"),
                before.get("kind"),
                before.get("category_hint"),
                before.get("transfer_group"),
                before.get("exchange_rate"),
                before.get("image_path"),
                before.get("commission_source_entry_id"),
                before.get("id"),
            ),
        )
        affected_expense_accounts.update(
            sync_statement_profit_entry(
                db,
                before.get("id"),
                enabled=bool(before.get("profit_expense_account_id")),
                account_id=before.get("profit_expense_account_id"),
            )
        )
    elif event["action"] == "delete":
        entry = payload.get("entry", {})
        db.execute(
            """
            insert into statement_entries (
                id, client_id, source_no, entry_date, description, currency, direction,
                amount, kind, category_hint, transfer_group, exchange_rate, image_path,
                commission_source_entry_id, profit_expense_entry_id, profit_expense_account_id
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                entry.get("id"),
                entry.get("client_id"),
                entry.get("source_no", 0),
                entry.get("entry_date"),
                entry.get("description"),
                entry.get("currency"),
                entry.get("direction"),
                entry.get("amount"),
                entry.get("kind"),
                entry.get("category_hint"),
                entry.get("transfer_group"),
                entry.get("exchange_rate"),
                entry.get("image_path"),
                entry.get("commission_source_entry_id"),
                entry.get("profit_expense_entry_id"),
                entry.get("profit_expense_account_id"),
            ),
        )
        affected_expense_accounts.update(
            sync_statement_profit_entry(
                db,
                entry.get("id"),
                enabled=bool(entry.get("profit_expense_account_id")),
                account_id=entry.get("profit_expense_account_id"),
            )
        )
    elif event["action"] == "exchange":
        for entry in payload.get("entries", []):
            db.execute("delete from statement_entries where id = ?", (entry.get("id"),))

    db.execute(
        "update statement_entry_events set undone_at = ? where id = ?",
        (utc_timestamp(), event["id"]),
    )
    db.commit()
    resequence_client_entries(client_id)
    for account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(account_id)
    return redirect(_return_url(client_id))


@app.route("/clients/<int:client_id>/print")
def print_statement(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    entries = db.execute(
        "select * from statement_entries where client_id = ? order by entry_date, id",
        (client_id,),
    ).fetchall()
    rows = statement_rows_with_commission_state(db, client_id, entries)
    return render_template("print.html", client=client, rows=rows)


def _attachment_response(payload: bytes | str, content_type: str, filename: str):
    response = make_response(payload)
    response.headers["Content-Type"] = content_type
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _statement_export_data(db: sqlite3.Connection, client_id: int) -> tuple[sqlite3.Row | None, list[dict]]:
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        return None, []
    entries = db.execute(
        f"select * from statement_entries where client_id = ? order by {ENTRY_ORDER}",
        (client_id,),
    ).fetchall()
    return client, statement_rows_with_commission_state(db, client_id, entries)


def _statement_pdf_response(client: sqlite3.Row, rows: list[dict]):
    final_usd = rows[-1]["running_usd"] if rows else 0.0
    final_cny = rows[-1]["running_cny"] if rows else 0.0
    html_string = render_template(
        "pdf.html",
        client=client,
        rows=rows,
        generated_date=china_today().isoformat(),
        final_usd=final_usd,
        final_cny=final_cny,
    )
    pdf_bytes = render_pdf_bytes(html_string)
    filename = f"{client['name'].replace(' ', '_')}_Statement.pdf"
    return _attachment_response(pdf_bytes, "application/pdf", filename)


def _statement_xlsx_response(client: sqlite3.Row, rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "personal Statment"

    header_font = Font(bold=True, size=14)
    col_font = Font(bold=True, size=10)
    col_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cny_symbol = currency_symbol("CNY")
    money_fmt_usd = '"$"#,##0.00'
    money_fmt_cny = f"{cny_symbol}#,##0.00"
    date_fmt = "YYYY-MM-DD"
    rate_fmt = "0.0000"
    neg_usd = '"$"#,##0.00;[Red]\\-"$"#,##0.00'
    neg_cny = f"{cny_symbol}#,##0.00;[Red]\\-{cny_symbol}#,##0.00"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"{client['name']} Statement"
    ws["A1"].font = header_font

    headers = [
        "No.",
        "Log",
        "Date",
        "In $",
        "Out $",
        f"in {cny_symbol}",
        f"out {cny_symbol}",
        "Balance $",
        f"Balance {cny_symbol}",
        "Rate in",
        "Rate out",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = col_font
        cell.fill = col_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    data_row = 3
    for row in rows:
        r = data_row
        amount = float(row["amount"])
        currency = row["currency"]
        direction = row["direction"]
        exchange_rate = row.get("exchange_rate")

        ws.cell(row=r, column=1, value=row["source_no"]).border = thin_border
        ws.cell(row=r, column=2, value=row["description"]).border = thin_border

        date_cell = ws.cell(row=r, column=3)
        date_cell.border = thin_border
        try:
            date_cell.value = datetime.strptime(row["entry_date"], "%Y-%m-%d")
            date_cell.number_format = date_fmt
        except (ValueError, TypeError):
            date_cell.value = row["entry_date"]

        in_usd_cell = ws.cell(row=r, column=4)
        in_usd_cell.border = thin_border
        if currency == "USD" and direction == "IN":
            in_usd_cell.value = amount
            in_usd_cell.number_format = money_fmt_usd

        out_usd_cell = ws.cell(row=r, column=5)
        out_usd_cell.border = thin_border
        if currency == "USD" and direction == "OUT":
            out_usd_cell.value = amount
            out_usd_cell.number_format = money_fmt_usd

        in_cny_cell = ws.cell(row=r, column=6)
        in_cny_cell.border = thin_border
        if currency == "CNY" and direction == "IN":
            in_cny_cell.value = amount
            in_cny_cell.number_format = money_fmt_cny

        out_cny_cell = ws.cell(row=r, column=7)
        out_cny_cell.border = thin_border
        if currency == "CNY" and direction == "OUT":
            out_cny_cell.value = amount
            out_cny_cell.number_format = money_fmt_cny

        bal_usd_cell = ws.cell(row=r, column=8, value=row["running_usd"])
        bal_usd_cell.border = thin_border
        bal_usd_cell.number_format = neg_usd

        bal_cny_cell = ws.cell(row=r, column=9, value=row["running_cny"])
        bal_cny_cell.border = thin_border
        bal_cny_cell.number_format = neg_cny

        rate_in_cell = ws.cell(row=r, column=10)
        rate_in_cell.border = thin_border
        if exchange_rate and direction == "IN":
            rate_in_cell.value = exchange_rate
            rate_in_cell.number_format = rate_fmt

        rate_out_cell = ws.cell(row=r, column=11)
        rate_out_cell.border = thin_border
        if exchange_rate and direction == "OUT":
            rate_out_cell.value = exchange_rate
            rate_out_cell.number_format = rate_fmt

        data_row += 1

    col_widths = [6, 45, 14, 14, 14, 14, 14, 16, 16, 12, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{client['name'].replace(' ', '_')}_Statement.xlsx"
    return _attachment_response(
        output.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename,
    )


def _expense_export_data(
    db: sqlite3.Connection, account_id: int
) -> tuple[sqlite3.Row | None, list[str], list[dict], dict[str, dict[str, float]]]:
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        return None, [], [], {}
    currencies = [c.strip() for c in account["enabled_currencies"].split(",") if c.strip()]
    entries = db.execute(
        "select * from expense_entries where account_id = ? order by entry_date, id",
        (account_id,),
    ).fetchall()
    rows = expense_running_balances(entries, currencies)
    totals: dict[str, dict[str, float]] = {}
    for currency in currencies:
        total_in = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='IN'",
            (account_id, currency),
        ).fetchone()[0]
        total_out = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='OUT'",
            (account_id, currency),
        ).fetchone()[0]
        totals[currency] = {"in": total_in, "out": total_out, "balance": total_in - total_out}
    return account, currencies, rows, totals


def _expense_xlsx_response(account: sqlite3.Row, currencies: list[str], rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    header_font = Font(bold=True, size=14)
    col_font = Font(bold=True, size=10)
    col_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    money_fmt = "#,##0.00"
    date_fmt = "YYYY-MM-DD"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(currencies) * 3)
    ws["A1"] = f"{account['name']} - Expense Report"
    ws["A1"].font = header_font

    headers = ["No.", "Date", "Description", "Category"]
    for currency in currencies:
        symbol = currency_symbol(currency)
        headers.extend([f"{symbol} In", f"{symbol} Out", f"{symbol} Balance"])

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = col_font
        cell.fill = col_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    data_row = 3
    for row in rows:
        r = data_row
        ws.cell(row=r, column=1, value=row["seq_no"]).border = thin_border
        date_cell = ws.cell(row=r, column=2)
        date_cell.border = thin_border
        try:
            date_cell.value = datetime.strptime(row["entry_date"], "%Y-%m-%d")
            date_cell.number_format = date_fmt
        except (ValueError, TypeError):
            date_cell.value = row["entry_date"]
        ws.cell(row=r, column=3, value=row["description"]).border = thin_border

        category_label = row["category"]
        for code, label in EXPENSE_CATEGORIES:
            if code == row["category"]:
                category_label = label
                break
        ws.cell(row=r, column=4, value=category_label).border = thin_border

        col_offset = 5
        for currency in currencies:
            in_cell = ws.cell(row=r, column=col_offset)
            in_cell.border = thin_border
            if row["currency"] == currency and row["direction"] == "IN":
                in_cell.value = float(row["amount"])
                in_cell.number_format = money_fmt

            out_cell = ws.cell(row=r, column=col_offset + 1)
            out_cell.border = thin_border
            if row["currency"] == currency and row["direction"] == "OUT":
                out_cell.value = float(row["amount"])
                out_cell.number_format = money_fmt

            bal_cell = ws.cell(row=r, column=col_offset + 2)
            bal_cell.border = thin_border
            bal_cell.value = row["running_balances"].get(currency, 0.0)
            bal_cell.number_format = money_fmt

            col_offset += 3

        data_row += 1

    col_widths = [6, 14, 40, 14] + [14, 14, 16] * len(currencies)
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{account['name'].replace(' ', '_')}_Expenses.xlsx"
    return _attachment_response(
        output.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename,
    )


def _expense_pdf_response(
    account: sqlite3.Row,
    currencies: list[str],
    rows: list[dict],
    totals: dict[str, dict[str, float]],
):
    html_string = render_template(
        "expense_pdf.html",
        account=account,
        currencies=currencies,
        rows=rows,
        totals=totals,
        generated_date=china_today().isoformat(),
        expense_categories=EXPENSE_CATEGORIES,
        currency_symbol=currency_symbol,
    )
    pdf_bytes = render_pdf_bytes(html_string)
    filename = f"{account['name'].replace(' ', '_')}_Expenses.pdf"
    return _attachment_response(pdf_bytes, "application/pdf", filename)


@app.route("/clients/<int:client_id>/export.pdf")
def export_pdf(client_id: int):
    db = get_db()
    client, rows = _statement_export_data(db, client_id)
    if client is None:
        abort(404)
    return _statement_pdf_response(client, rows)


@app.route("/clients/<int:client_id>/export.csv")
def export_statement(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    entries = db.execute(
        """
        select
            entry_date,
            source_no,
            description,
            currency,
            direction,
            amount,
            kind,
            category_hint,
            transfer_group,
            commission_source_entry_id,
            profit_expense_entry_id,
            profit_expense_account_id
        from statement_entries
        where client_id = ?
        order by entry_date, id
        """,
        (client_id,),
    ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "date",
            "source_no",
            "description",
            "currency",
            "direction",
            "amount",
            "kind",
            "category_hint",
            "transfer_group",
            "commission_source_entry_id",
            "profit_expense_entry_id",
            "profit_expense_account_id",
        ]
    )
    for entry in entries:
        writer.writerow(
            [
                entry["entry_date"],
                entry["source_no"],
                entry["description"],
                entry["currency"],
                entry["direction"],
                entry["amount"],
                entry["kind"],
                entry["category_hint"],
                entry["transfer_group"],
                entry["commission_source_entry_id"],
                entry["profit_expense_entry_id"],
                entry["profit_expense_account_id"],
            ]
        )
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{client["name"].replace(" ", "_").lower()}_statement.csv"'
    return response


@app.route("/clients/<int:client_id>/export.xlsx")
def export_statement_xlsx(client_id: int):
    db = get_db()
    client, rows = _statement_export_data(db, client_id)
    if client is None:
        abort(404)
    return _statement_xlsx_response(client, rows)


@app.route("/clients/<int:client_id>/delete", methods=["POST"])
@admin_required
def delete_client(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    db.execute("delete from statement_entry_events where client_id = ?", (client_id,))
    db.execute("delete from statement_entries where client_id = ?", (client_id,))
    db.execute("delete from clients where id = ?", (client_id,))
    db.commit()
    return redirect(url_for("index"))


@app.route("/clients/group", methods=["POST"])
def group_clients():
    """Group selected clients under a parent. The parent keeps its own entries separate."""
    parent_id_raw = request.form.get("parent_id", "").strip()
    child_ids_raw = request.form.getlist("child_ids")
    if not parent_id_raw or not child_ids_raw:
        return render_index("Select a parent and at least one client to group under it."), 400

    parent_id = int(parent_id_raw)
    child_ids = [int(cid) for cid in child_ids_raw if cid != parent_id_raw]
    if not child_ids:
        return render_index("Children must be different from the parent."), 400

    db = get_db()
    parent = db.execute("select * from clients where id = ?", (parent_id,)).fetchone()
    if parent is None:
        return render_index("Parent client not found."), 400

    for cid in child_ids:
        child = db.execute("select * from clients where id = ?", (cid,)).fetchone()
        if child is None:
            continue
        db.execute("update clients set parent_id = ? where id = ?", (parent_id, cid))

    db.commit()
    return redirect(url_for("index"))


@app.route("/clients/<int:client_id>/ungroup", methods=["POST"])
def ungroup_client(client_id: int):
    """Remove a client from its parent group (make standalone)."""
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    db.execute("update clients set parent_id = null where id = ?", (client_id,))
    db.commit()
    return redirect(url_for("index"))


@app.route("/clients/<int:client_id>/ungroup-all", methods=["POST"])
def ungroup_all_children(client_id: int):
    """Ungroup all children from this parent."""
    db = get_db()
    db.execute("update clients set parent_id = null where parent_id = ?", (client_id,))
    db.commit()
    return redirect(url_for("index"))


@app.route("/clients/<int:client_id>/rename", methods=["POST"])
def rename_client(client_id: int):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        abort(404)
    new_name = request.form.get("name", "").strip()
    if not new_name:
        return redirect(url_for("index"))
    existing = db.execute("select id from clients where name = ? and id != ?", (new_name, client_id)).fetchone()
    if existing is not None:
        return render_index(f'Client name "{new_name}" is already taken.'), 400
    db.execute("update clients set name = ? where id = ?", (new_name, client_id))
    db.commit()
    return redirect(url_for("index"))


@app.route("/bank-balance/add", methods=["POST"])
def add_bank_balance():
    account_name = request.form.get("account_name", "").strip()
    if not account_name:
        return render_index("Account name is required."), 400
    usd = float(request.form.get("usd_balance", 0) or 0)
    cny = float(request.form.get("cny_balance", 0) or 0)
    db = get_db()
    db.execute(
        "insert into bank_balances (account_name, usd_balance, cny_balance, updated_at) values (?, ?, ?, ?)",
        (account_name, usd, cny, utc_timestamp()),
    )
    db.commit()
    return redirect(url_for("index"))


@app.route("/bank-balance/<int:balance_id>/edit", methods=["POST"])
def edit_bank_balance(balance_id: int):
    db = get_db()
    row = db.execute("select * from bank_balances where id = ?", (balance_id,)).fetchone()
    if row is None:
        abort(404)
    account_name = request.form.get("account_name", "").strip() or row["account_name"]
    usd = float(request.form.get("usd_balance", 0) or 0)
    cny = float(request.form.get("cny_balance", 0) or 0)
    db.execute(
        "update bank_balances set account_name = ?, usd_balance = ?, cny_balance = ?, updated_at = ? where id = ?",
        (account_name, usd, cny, utc_timestamp(), balance_id),
    )
    db.commit()
    return redirect(url_for("index"))


@app.route("/bank-balance/<int:balance_id>/delete", methods=["POST"])
def delete_bank_balance(balance_id: int):
    db = get_db()
    db.execute("delete from bank_balances where id = ?", (balance_id,))
    db.commit()
    return redirect(url_for("index"))


@app.route("/supplier/add", methods=["POST"])
def add_supplier():
    supplier_name = request.form.get("supplier_name", "").strip()
    if not supplier_name:
        return render_index("Supplier name is required."), 400
    currency = request.form.get("currency", "CNY").strip().upper()
    if currency not in ("USD", "CNY"):
        currency = "CNY"
    amount = float(request.form.get("amount_owed", 0) or 0)
    notes = request.form.get("notes", "").strip()
    db = get_db()
    db.execute(
        "insert into supplier_balances (supplier_name, currency, amount_owed, notes, updated_at) values (?, ?, ?, ?, ?)",
        (supplier_name, currency, amount, notes, utc_timestamp()),
    )
    db.commit()
    return redirect(url_for("index"))


@app.route("/supplier/<int:supplier_id>/edit", methods=["POST"])
def edit_supplier(supplier_id: int):
    db = get_db()
    row = db.execute("select * from supplier_balances where id = ?", (supplier_id,)).fetchone()
    if row is None:
        abort(404)
    supplier_name = request.form.get("supplier_name", "").strip() or row["supplier_name"]
    currency = request.form.get("currency", "CNY").strip().upper()
    if currency not in ("USD", "CNY"):
        currency = row["currency"]
    amount = float(request.form.get("amount_owed", 0) or 0)
    notes = request.form.get("notes", "").strip()
    db.execute(
        "update supplier_balances set supplier_name = ?, currency = ?, amount_owed = ?, notes = ?, updated_at = ? where id = ?",
        (supplier_name, currency, amount, notes, utc_timestamp(), supplier_id),
    )
    db.commit()
    return redirect(url_for("index"))


@app.route("/supplier/<int:supplier_id>/delete", methods=["POST"])
def delete_supplier(supplier_id: int):
    db = get_db()
    db.execute("delete from supplier_balances where id = ?", (supplier_id,))
    db.commit()
    return redirect(url_for("index"))


@app.route("/uploads/<path:filename>")
def serve_upload(filename: str):
    return send_from_directory(UPLOAD_DIR, filename)


# ---------------------------------------------------------------------------
# Quick Submit — upload images on the go, process into statements later
# ---------------------------------------------------------------------------

@app.route("/quick-submit")
def quick_submit_page():
    pending = get_db().execute("select count(*) from quick_submits where status = 'pending'").fetchone()[0]
    return render_template("quick_submit.html", clients=client_list(), pending_count=pending)


@app.route("/quick-submit", methods=["POST"])
def quick_submit_save():
    image_file = request.files.get("image")
    if not image_file or not image_file.filename:
        flash("Image is required", "error")
        return redirect(url_for("quick_submit_page"))
    client_id = request.form.get("client_id", type=int)
    if not client_id:
        flash("Please select a client", "error")
        return redirect(url_for("quick_submit_page"))
    image_filename = save_upload_image(image_file)
    if not image_filename:
        flash("Invalid image type", "error")
        return redirect(url_for("quick_submit_page"))
    amount_raw = request.form.get("amount", "").strip()
    amount = float(amount_raw) if amount_raw else None
    db = get_db()
    db.execute(
        "insert into quick_submits (client_id, description, amount, image_path, status, created_at, created_by) values (?, ?, ?, ?, 'pending', ?, ?)",
        (client_id, request.form.get("description", "").strip(), amount, image_filename, utc_timestamp(), g.user["username"]),
    )
    db.commit()
    flash("Image submitted successfully", "success")
    return redirect(url_for("quick_submit_page"))


@app.route("/quick-submit/history")
def quick_submit_history():
    db = get_db()
    rows = db.execute(
        "select qs.*, c.name as client_name from quick_submits qs left join clients c on c.id = qs.client_id order by qs.created_at desc"
    ).fetchall()
    return render_template("quick_submit_history.html", submissions=rows)


@app.route("/quick-submit/<int:submit_id>/process", methods=["POST"])
def quick_submit_process(submit_id: int):
    db = get_db()
    row = db.execute("select * from quick_submits where id = ?", (submit_id,)).fetchone()
    if row is None:
        abort(404)
    return redirect(url_for("client_statement", client_id=row["client_id"], quick_submit_id=submit_id))


@app.route("/quick-submit/<int:submit_id>/delete", methods=["POST"])
def quick_submit_delete(submit_id: int):
    db = get_db()
    row = db.execute("select * from quick_submits where id = ?", (submit_id,)).fetchone()
    if row is None:
        abort(404)
    _delete_image_file(row["image_path"])
    db.execute("delete from quick_submits where id = ?", (submit_id,))
    db.commit()
    flash("Submission deleted", "success")
    return redirect(url_for("quick_submit_history"))


PARSE_IMAGE_PROMPT = (
    "You extract transaction data from receipt/invoice/transfer images. "
    "Return ONLY a JSON object with these fields (use null for unknown): "
    '{"date": "YYYY-MM-DD", "description": "...", '
    '"currency": "USD or CNY", "direction": "IN or OUT", '
    '"amount": 123.45, "category": "...", '
    '"recipient_name": "...", "card_last4": "1234", '
    '"extracted_text": "..."}\n'
    "Rules:\n"
    "- date: Extract the exact date from the image. Convert any date format to YYYY-MM-DD. "
    "If only month/year is shown, use the 1st of that month. Never return null if any date is visible.\n"
    "- description: A concise one-line summary of the transaction.\n"
    "- direction: OUT if you are sending/paying/transferring money to someone. "
    "IN if you are receiving money. Transfers and payments are always OUT.\n"
    "- recipient_name: IMPORTANT - Always extract the recipient/payee name from the image. "
    "Look for: 收款人, 对方户名, 收款方, payee, recipient, transfer to, etc. "
    "If the name is in Chinese characters, return as 'Pinyin Name (中文名)' "
    "e.g. 'Zhang San (张三)'. If the name is already in English, keep it as-is in English only — do NOT add Chinese translation. "
    "null ONLY if absolutely no name is visible anywhere in the image.\n"
    "- card_last4: Extract the last 4 digits of any bank card number visible in the image. null if not visible.\n"
    "- category: one of: client_receipt, factory_payment, commission, shipping_expense, "
    "travel_expense, internal_transfer, fx_transfer, uncategorized\n"
    "- extracted_text: ALL visible text from the image, organized by sections using markdown "
    "(## for headers). Preserve original Chinese characters. Do NOT translate. "
    "Include every detail: names, numbers, dates, amounts, bank info, notes, etc."
)

EXTRACT_TEXT_PROMPT = (
    "You are an OCR assistant. Extract ALL text visible in the image and return it "
    "in a clean, well-structured format. Organize the text logically by sections as they appear "
    "in the document (e.g. header, sender info, recipient info, transaction details, notes, footer). "
    "Use clear section headers with markdown formatting (## for sections). "
    "Preserve important details like names, numbers, dates, and amounts exactly as shown. "
    "For Chinese text, keep the original Chinese characters. "
    "Do NOT translate anything. Do NOT add information that is not in the image. "
    "Do NOT wrap in code blocks. Return clean readable text only."
)


@app.route("/api/parse-image", methods=["POST"])
def parse_image_api():
    """Parse an uploaded image using OpenRouter to extract transaction details."""
    try:
        image_data, mime = _resolve_image_upload(file=request.files.get("image"), image_url=request.form.get("image_url"))
        content = _call_vision(image_data, mime, PARSE_IMAGE_PROMPT)
        json_match = re.search(r"\{.*\}", content, re.DOTALL)
        if json_match:
            parsed = json.loads(json_match.group())
            return jsonify({"ok": True, "data": parsed})
        return jsonify({"ok": True, "data": {"raw": content}})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/extract-text", methods=["POST"])
def api_extract_text():
    try:
        image_data, mime = _resolve_image_upload(
            file=request.files.get("image"),
            image_url=request.form.get("image_url"),
        )
        content = _call_vision(image_data, mime, EXTRACT_TEXT_PROMPT,
                                     user_prompt="Extract and structure all visible text from this image.")
        return jsonify({"ok": True, "text": content})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/link-transfer", methods=["POST"])
@admin_required
def api_link_transfer():
    data = request.get_json()
    entry_id_1 = data.get("entry_id_1")
    entry_id_2 = data.get("entry_id_2")
    if not entry_id_1 or not entry_id_2:
        return jsonify({"error": "Missing entry IDs"}), 400

    db = get_db()
    selected_entries = db.execute(
        "select id, commission_source_entry_id from statement_entries where id in (?, ?)",
        (entry_id_1, entry_id_2),
    ).fetchall()
    if len(selected_entries) != 2:
        return jsonify({"error": "One or more entries no longer exist"}), 404
    if any(row["commission_source_entry_id"] for row in selected_entries):
        return jsonify({"error": "Generated commission rows cannot be linked into an FX transfer."}), 409
    group = make_transfer_group()
    db.execute("UPDATE statement_entries SET transfer_group = ? WHERE id = ?", (group, entry_id_1))
    db.execute("UPDATE statement_entries SET transfer_group = ? WHERE id = ?", (group, entry_id_2))
    db.commit()
    sync_exchange_group(db, group)
    db.commit()

    # Return updated entries with exchange_rate
    def entry_dict(eid):
        row = db.execute("select * from statement_entries where id = ?", (eid,)).fetchone()
        if not row:
            return None
        return {
            "id": row["id"],
            "exchange_rate": row["exchange_rate"],
            "kind": row["kind"],
            "transfer_group": row["transfer_group"] or "",
            "linked_entry_id": row["linked_entry_id"],
        }

    return jsonify({
        "ok": True,
        "transfer_group": group,
        "entries": [entry_dict(entry_id_1), entry_dict(entry_id_2)],
    })


# ---------------------------------------------------------------------------
# FX Rate API — moneyconvert.net → open.er-api.com → DB fallback
# ---------------------------------------------------------------------------
_fx_cache: dict = {}  # keyed by base currency e.g. "USD" -> {rates: {...}, fetched_at: float, source: str}
_FX_CACHE_TTL = 3600  # 1 hour

MONEYCONVERT_URL = "https://cdn.moneyconvert.net/api/latest.json"
ER_API_URL = "https://open.er-api.com/v6/latest/{base}"
FX_PROXY_URL = os.environ.get("FX_PROXY_URL", "").strip()


def _fetch_moneyconvert(proxies):
    """Fetch from moneyconvert.net (hourly updates, all rates based on USD)."""
    with http_requests.Session() as session:
        session.trust_env = False
        resp = session.get(MONEYCONVERT_URL, proxies=proxies, timeout=8)
    resp.raise_for_status()
    data = resp.json()
    rates = data.get("rates")
    if rates and isinstance(rates, dict):
        return rates
    return None


def _fetch_er_api(base, proxies):
    """Fetch from open.er-api.com (daily updates, any base currency)."""
    with http_requests.Session() as session:
        session.trust_env = False
        resp = session.get(ER_API_URL.format(base=base), proxies=proxies, timeout=8)
    resp.raise_for_status()
    data = resp.json()
    if data.get("result") == "success":
        return data["rates"]
    return None


def _convert_rates_from_usd(usd_rates, base):
    """Convert USD-based rates dict to rates relative to a different base."""
    base_in_usd = usd_rates.get(base)
    if not base_in_usd or base_in_usd == 0:
        return None
    return {cur: val / base_in_usd for cur, val in usd_rates.items()}


@app.route("/api/fx-rate")
def api_fx_rate():
    """Return exchange rate for a currency pair (no login required)."""
    import time

    from_cur = request.args.get("from", "USD").upper().strip()
    to_cur = request.args.get("to", "CNY").upper().strip()

    if from_cur == "RMB":
        from_cur = "CNY"
    if to_cur == "RMB":
        to_cur = "CNY"

    if from_cur == to_cur:
        ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")
        return jsonify({"rate": 1.0, "from": from_cur, "to": to_cur, "source": "identity", "timestamp": ts})

    now = time.time()

    # Check cache
    cached = _fx_cache.get(from_cur)
    if cached and (now - cached["fetched_at"]) < _FX_CACHE_TTL:
        rate = cached["rates"].get(to_cur)
        if rate is not None:
            return jsonify({"rate": rate, "from": from_cur, "to": to_cur, "source": cached["source"], "timestamp": cached["timestamp"]})

    proxies = {"http": FX_PROXY_URL, "https": FX_PROXY_URL} if FX_PROXY_URL else None
    ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")

    # 1) Try moneyconvert.net (hourly, USD-based)
    try:
        usd_rates = _fetch_moneyconvert(proxies)
        if usd_rates:
            if from_cur == "USD":
                rates = usd_rates
            else:
                rates = _convert_rates_from_usd(usd_rates, from_cur)
            if rates:
                _fx_cache[from_cur] = {"rates": rates, "fetched_at": now, "timestamp": ts, "source": "moneyconvert"}
                rate = rates.get(to_cur)
                if rate is not None:
                    return jsonify({"rate": rate, "from": from_cur, "to": to_cur, "source": "moneyconvert", "timestamp": ts})
    except Exception:
        pass

    # 2) Try open.er-api.com (daily, any base)
    try:
        rates = _fetch_er_api(from_cur, proxies)
        if rates:
            _fx_cache[from_cur] = {"rates": rates, "fetched_at": now, "timestamp": ts, "source": "exchangerate-api"}
            rate = rates.get(to_cur)
            if rate is not None:
                return jsonify({"rate": rate, "from": from_cur, "to": to_cur, "source": "exchangerate-api", "timestamp": ts})
    except Exception:
        pass

    # 3) Fallback: stored rate from database (USD/CNY only)
    if (from_cur == "USD" and to_cur == "CNY") or (from_cur == "CNY" and to_cur == "USD"):
        try:
            summary = exchange_rate_summary()
            stored_rate = summary.get("latest_rate")
            if stored_rate is not None:
                rate = stored_rate if from_cur == "USD" else 1.0 / stored_rate
                return jsonify({"rate": rate, "from": from_cur, "to": to_cur, "source": "stored", "timestamp": ts})
        except Exception:
            pass

    return jsonify({"error": "Unable to fetch exchange rate for " + from_cur + "/" + to_cur}), 503


@app.route("/api/global-search")
def global_search():
    """Search entries across all clients by any field."""
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify([])
    db = get_db()
    q_pat = f"%{q}%"
    results = db.execute(
        """select se.id, se.client_id, c.name as client_name, se.entry_date,
                  se.description, se.currency, se.direction, se.amount,
                  se.source_no, se.kind, se.category_hint
           from statement_entries se join clients c on se.client_id = c.id
           where se.description like ?
              or cast(se.amount as text) like ?
              or se.entry_date like ?
              or se.currency like ?
              or se.kind like ?
              or coalesce(se.category_hint,'') like ?
              or c.name like ?
              or coalesce(se.transfer_group,'') like ?
           order by se.entry_date desc
           limit 20""",
        (q_pat, q_pat, q_pat, q_pat, q_pat, q_pat, q_pat, q_pat),
    ).fetchall()
    return jsonify([dict(r) for r in results])


@app.route("/api/fx-refresh", methods=["POST"])
@admin_required
def api_fx_refresh():
    """Fetch live USD→CNY rate from API and store it in app_settings."""
    rate = _get_live_usd_cny_rate()
    if rate is None:
        return jsonify({"error": "Unable to fetch live exchange rate"}), 503
    ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")
    set_setting("fx_live_rate", str(rate))
    set_setting("fx_live_rate_updated_at", ts)
    # Determine source from cache
    cached = _fx_cache.get("USD")
    source = cached.get("source", "api") if cached else "api"
    return jsonify({"rate": rate, "source": source, "updated_at": ts})


@app.route("/settings", methods=["GET"])
@admin_required
def settings_page():
    # Mask the API key for display
    raw_key = get_openrouter_api_key()
    masked_key = raw_key[:6] + "..." + raw_key[-4:] if len(raw_key) > 10 else raw_key
    # Fetch API audit log
    db = get_db()
    audit_rows = db.execute("select * from api_audit_log order by id desc limit 10").fetchall()
    audit_log = []
    for r in audit_rows:
        d = dict(r)
        d["detail"] = json.loads(d["detail"]) if d["detail"] else {}
        audit_log.append(d)
    return render_template(
        "settings.html",
        api_key_masked=masked_key,
        api_key_set=bool(raw_key),
        current_model=get_openrouter_model(),
        models=_fetch_openrouter_models(),
        audit_log=audit_log,
        current_db_info=database_file_info(DB_PATH),
        db_backups=list_database_backups(),
        db_path_display=str(DB_PATH),
        backup_dir_display=str(BACKUP_DIR),
        fx_rate_source=get_setting("fx_rate_source", "live"),
        fx_live_rate=get_setting("fx_live_rate", ""),
        fx_live_rate_updated_at=get_setting("fx_live_rate_updated_at", ""),
    )


@app.route("/settings", methods=["POST"])
@admin_required
def settings_save():
    db = get_db()
    api_key = request.form.get("openrouter_api_key", "").strip()
    model = request.form.get("openrouter_model", "").strip()
    fx_source = request.form.get("fx_rate_source", "").strip()
    if api_key:
        db.execute("insert or replace into app_settings(key, value) values (?, ?)", ("openrouter_api_key", api_key))
    if model:
        db.execute("insert or replace into app_settings(key, value) values (?, ?)", ("openrouter_model", model))
    if fx_source in ("live", "average"):
        db.execute("insert or replace into app_settings(key, value) values (?, ?)", ("fx_rate_source", fx_source))
    db.commit()
    flash("Settings saved", "success")
    return redirect(url_for("settings_page"))


@app.route("/settings/database/download")
@admin_required
def settings_download_database():
    ensure_runtime_dirs()
    with tempfile.NamedTemporaryFile(delete=False, dir=DATA_DIR, prefix="download-", suffix=".db") as tmp:
        snapshot_path = Path(tmp.name)
    snapshot_database(snapshot_path)

    @after_this_request
    def cleanup_snapshot(response):
        snapshot_path.unlink(missing_ok=True)
        return response

    download_name = f"{DB_PATH.stem}-{utc_now().strftime('%Y%m%dT%H%M%SZ')}.db"
    return send_file(snapshot_path, as_attachment=True, download_name=download_name)


@app.route("/settings/backup/download")
@admin_required
def settings_download_full_backup():
    ensure_runtime_dirs()
    backup_path = DATA_DIR / f"download-{uuid.uuid4().hex}.tar.gz"
    create_full_backup(backup_path)

    @after_this_request
    def cleanup_backup(response):
        backup_path.unlink(missing_ok=True)
        return response

    return send_file(
        backup_path,
        as_attachment=True,
        download_name=make_full_backup_name(),
        mimetype="application/gzip",
    )


@app.route("/settings/database/upload", methods=["POST"])
@admin_required
def settings_upload_database():
    cleanup_dir = None
    try:
        candidate_path, staged_uploads, original_name, restore_kind, cleanup_dir = stage_uploaded_restore(
            request.files.get("database_file")
        )
        previous_backup = activate_restore_candidate(candidate_path, f"upload-{original_name}", staged_uploads)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("settings_page"))
    except Exception as exc:
        flash(f"Restore failed: {exc}", "error")
        return redirect(url_for("settings_page"))
    finally:
        if cleanup_dir is not None:
            shutil.rmtree(cleanup_dir, ignore_errors=True)
    session.clear()
    restore_label = "Full backup" if restore_kind == "full" else "Database-only file"
    flash(
        f"{restore_label} restored from '{original_name}'. Previous full backup saved as '{previous_backup.name}'. Sign in again.",
        "success",
    )
    return redirect(url_for("login_page"))


@app.route("/settings/database/backups/<path:backup_name>/download")
@admin_required
def settings_download_database_backup(backup_name: str):
    backup_path = resolve_backup_path(backup_name)
    return send_file(backup_path, as_attachment=True, download_name=backup_path.name)


@app.route("/settings/database/backups/restore", methods=["POST"])
@admin_required
def settings_restore_database_backup():
    backup_name = request.form.get("backup_name", "").strip()
    if not backup_name:
        flash("Choose a backup to restore.", "error")
        return redirect(url_for("settings_page"))
    cleanup_dir = None
    try:
        candidate_path, staged_uploads, staged_name, restore_kind, cleanup_dir = stage_backup_restore(backup_name)
        previous_backup = activate_restore_candidate(candidate_path, f"restore-{staged_name}", staged_uploads)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("settings_page"))
    except Exception as exc:
        flash(f"Backup restore failed: {exc}", "error")
        return redirect(url_for("settings_page"))
    finally:
        if cleanup_dir is not None:
            shutil.rmtree(cleanup_dir, ignore_errors=True)
    session.clear()
    restore_label = "Full backup" if restore_kind == "full" else "Database-only backup"
    flash(
        f"{restore_label} '{staged_name}' restored. Previous full backup saved as '{previous_backup.name}'. Sign in again.",
        "success",
    )
    return redirect(url_for("login_page"))


@app.route("/reload", methods=["POST"])
@admin_required
def reload_data():
    init_db()
    if seed_from_csv():
        flash("Demo data reloaded.", "success")
    else:
        flash("Demo CSV is not configured. Set SOURCE_CSV_PATH to import demo data.", "error")
    return redirect(url_for("index"))


# --- Expense routes ---

@app.route("/expenses/")
def expense_list():
    accounts = expense_dashboard_data()
    return render_template(
        "expenses.html",
        accounts=accounts,
        all_currencies=ALL_EXPENSE_CURRENCIES,
    )


@app.route("/expenses/accounts/new", methods=["POST"])
def expense_account_new():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Account name is required", "error")
        return redirect(url_for("expense_list"))
    currencies = request.form.getlist("currencies")
    if not currencies:
        currencies = ["CNY"]
    valid = [c for c in currencies if c in ALL_EXPENSE_CURRENCIES]
    if not valid:
        valid = ["CNY"]
    db = get_db()
    existing = db.execute("select id from expense_accounts where name = ?", (name,)).fetchone()
    if existing:
        flash(f"Account '{name}' already exists", "error")
        return redirect(url_for("expense_list"))
    db.execute(
        "insert into expense_accounts (name, enabled_currencies, created_at) values (?, ?, ?)",
        (name, ",".join(valid), utc_timestamp()),
    )
    db.commit()
    account = db.execute("select id from expense_accounts where name = ?", (name,)).fetchone()
    return redirect(url_for("expense_account_view", account_id=account["id"]))


@app.route("/expenses/accounts/<int:account_id>")
def expense_account_view(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    currencies = [c.strip() for c in account["enabled_currencies"].split(",") if c.strip()]

    # Generate recurring entries
    generate_recurring_expenses(account_id)

    # Filters
    filters = {
        "q": request.args.get("q", "").strip(),
        "category": request.args.get("category", "").strip(),
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
    }

    clauses = ["account_id = ?"]
    params = [account_id]
    if filters["q"]:
        clauses.append("description like ?")
        params.append(f'%{filters["q"]}%')
    if filters["category"]:
        clauses.append("category = ?")
        params.append(filters["category"])
    if filters["date_from"]:
        clauses.append("entry_date >= ?")
        params.append(filters["date_from"])
    if filters["date_to"]:
        clauses.append("entry_date <= ?")
        params.append(filters["date_to"])

    where_sql = " and ".join(clauses)
    all_entries = db.execute(
        f"select * from expense_entries where {where_sql} order by entry_date, id",
        params,
    ).fetchall()

    rows = annotate_expense_rows(db, expense_running_balances(all_entries, currencies))

    # Totals
    totals = {}
    for cur in currencies:
        total_in = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='IN'",
            (account_id, cur),
        ).fetchone()[0]
        total_out = db.execute(
            "select coalesce(sum(amount),0) from expense_entries where account_id=? and currency=? and direction='OUT'",
            (account_id, cur),
        ).fetchone()[0]
        totals[cur] = {"in": total_in, "out": total_out, "balance": total_in - total_out}

    # Latest undo event
    latest_event = db.execute(
        "select * from expense_events where account_id = ? and undone_at is null order by id desc limit 1",
        (account_id,),
    ).fetchone()

    # Templates
    templates = db.execute(
        "select * from recurring_expense_templates where account_id = ? order by description",
        (account_id,),
    ).fetchall()

    return render_template(
        "expense_account.html",
        account=account,
        currencies=currencies,
        all_currencies=ALL_EXPENSE_CURRENCIES,
        rows=rows,
        totals=totals,
        latest_event=latest_event,
        templates=templates,
        filters=filters,
        default_date=china_today().isoformat(),
        expense_categories=EXPENSE_CATEGORIES,
        currency_symbol=currency_symbol,
    )


@app.route("/expenses/accounts/<int:account_id>/settings", methods=["POST"])
def expense_account_settings(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    currencies = request.form.getlist("currencies")
    valid = [c for c in currencies if c in ALL_EXPENSE_CURRENCIES]
    if not valid:
        valid = ["CNY"]
    db.execute(
        "update expense_accounts set enabled_currencies = ? where id = ?",
        (",".join(valid), account_id),
    )
    db.commit()
    return redirect(url_for("expense_account_view", account_id=account_id))


@app.route("/expenses/accounts/<int:account_id>/rename", methods=["POST"])
def expense_account_rename(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    new_name = request.form.get("name", "").strip()
    if not new_name:
        return redirect(url_for("expense_account_view", account_id=account_id))
    existing = db.execute("select id from expense_accounts where name = ? and id != ?", (new_name, account_id)).fetchone()
    if existing:
        flash(f"Account name '{new_name}' is already taken", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))
    db.execute("update expense_accounts set name = ? where id = ?", (new_name, account_id))
    db.commit()
    return redirect(url_for("expense_account_view", account_id=account_id))


@app.route("/expenses/accounts/<int:account_id>/delete", methods=["POST"])
@admin_required
def expense_account_delete(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    linked_count = db.execute(
        "select count(*) from expense_entries where account_id = ? and linked_statement_entry_id is not null",
        (account_id,),
    ).fetchone()[0]
    if linked_count:
        flash("This account contains statement-linked commission entries. Unlink them from the client statement first.", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))
    db.execute("delete from expense_events where account_id = ?", (account_id,))
    db.execute("delete from expense_entries where account_id = ?", (account_id,))
    db.execute("delete from recurring_expense_templates where account_id = ?", (account_id,))
    db.execute("delete from expense_accounts where id = ?", (account_id,))
    db.commit()
    return redirect(url_for("expense_list"))


@app.route("/expenses/accounts/<int:account_id>/entries", methods=["POST"])
def expense_add_entry(account_id):
    db = get_db()
    account = db.execute("select id from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    image_filename = save_upload_image(request.files.get("image"))
    entry_date = request.form["entry_date"]
    description = request.form["description"].strip()
    currency = request.form["currency"]
    direction = request.form.get("direction", "OUT")
    amount = float(request.form["amount"])
    category = request.form.get("category", "general")

    db.execute(
        """insert into expense_entries
        (account_id, seq_no, entry_date, description, currency, direction,
         amount, category, is_recurring, template_id, image_path, created_at)
        values (?, 0, ?, ?, ?, ?, ?, ?, 0, null, ?, ?)""",
        (account_id, entry_date, description, currency, direction, amount, category, image_filename, utc_timestamp()),
    )
    db.commit()
    entry = db.execute(
        "select * from expense_entries where account_id = ? order by id desc limit 1",
        (account_id,),
    ).fetchone()
    if entry:
        expense_record_event(account_id, entry["id"], "add", {"entry": expense_entry_to_dict(entry)})
    resequence_expense_entries(account_id)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        fresh = db.execute("select * from expense_entries where id = ?", (entry["id"],)).fetchone()
        return jsonify({
            "ok": True,
            "entry": {
                "id": fresh["id"],
                "entry_date": fresh["entry_date"],
                "seq_no": fresh["seq_no"],
                "description": fresh["description"],
                "currency": fresh["currency"],
                "direction": fresh["direction"],
                "amount": float(fresh["amount"]),
                "category": fresh["category"],
                "linked_statement_entry_id": fresh["linked_statement_entry_id"],
                "image_path": fresh["image_path"] or "",
            },
        })
    return redirect(url_for("expense_account_view", account_id=account_id) + "#table-top")


@app.route("/expenses/entries/<int:entry_id>/save", methods=["POST"])
def expense_save_entry(entry_id):
    db = get_db()
    entry = db.execute("select * from expense_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        abort(404)
    if entry["linked_statement_entry_id"]:
        message = linked_expense_entry_message(db, entry)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": message}), 409
        flash(message, "error")
        return redirect(url_for("expense_account_view", account_id=entry["account_id"]) + f"#entry-{entry_id}")
    before = expense_entry_to_dict(entry)
    old_image = entry["image_path"]
    new_image = save_upload_image(request.files.get("image"))
    image_path = new_image if new_image else old_image
    if request.form.get("remove_image") == "1":
        image_path = None
    if old_image and old_image != image_path:
        _delete_image_file(old_image)

    db.execute(
        """update expense_entries
        set entry_date = ?, description = ?, currency = ?, direction = ?,
            amount = ?, category = ?, image_path = ?
        where id = ?""",
        (
            request.form["entry_date"],
            request.form["description"].strip(),
            request.form["currency"],
            request.form.get("direction", "OUT"),
            float(request.form["amount"]),
            request.form.get("category", "general"),
            image_path,
            entry_id,
        ),
    )
    db.commit()
    updated = db.execute("select * from expense_entries where id = ?", (entry_id,)).fetchone()
    if updated:
        expense_record_event(entry["account_id"], entry_id, "edit", {"before": before, "after": expense_entry_to_dict(updated)})
    resequence_expense_entries(entry["account_id"])

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        updated_row = db.execute("select * from expense_entries where id = ?", (entry_id,)).fetchone()
        return jsonify({
            "ok": True,
            "entry": {
                "id": entry_id,
                "entry_date": updated_row["entry_date"],
                "seq_no": updated_row["seq_no"],
                "description": updated_row["description"],
                "currency": updated_row["currency"],
                "direction": updated_row["direction"],
                "amount": updated_row["amount"],
                "category": updated_row["category"],
                "linked_statement_entry_id": updated_row["linked_statement_entry_id"],
                "image_path": updated_row["image_path"] or "",
            },
        })
    return redirect(url_for("expense_account_view", account_id=entry["account_id"]) + f"#entry-{entry_id}")


@app.route("/expenses/entries/<int:entry_id>/delete", methods=["POST"])
def expense_delete_entry(entry_id):
    db = get_db()
    entry = db.execute("select * from expense_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        abort(404)
    if entry["linked_statement_entry_id"]:
        message = linked_expense_entry_message(db, entry)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": message}), 409
        flash(message, "error")
        return redirect(url_for("expense_account_view", account_id=entry["account_id"]) + f"#entry-{entry_id}")
    expense_record_event(entry["account_id"], entry_id, "delete", {"entry": expense_entry_to_dict(entry)})
    _delete_image_file(entry["image_path"])
    db.execute("delete from expense_entries where id = ?", (entry_id,))
    db.commit()
    resequence_expense_entries(entry["account_id"])
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "deleted_id": entry_id})
    return redirect(url_for("expense_account_view", account_id=entry["account_id"]) + "#table-top")


@app.route("/expenses/accounts/<int:account_id>/balances")
def expense_account_balances(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    currencies = [c.strip() for c in account["enabled_currencies"].split(",") if c.strip()]
    all_entries = db.execute(
        "select * from expense_entries where account_id = ? order by entry_date, id",
        (account_id,),
    ).fetchall()
    rows = expense_running_balances(all_entries, currencies)
    balances = {}
    for row in rows:
        balances[str(row["id"])] = {"running_balances": row["running_balances"], "seq_no": row["seq_no"]}
    return jsonify({"ok": True, "balances": balances, "currencies": currencies})


@app.route("/expenses/accounts/<int:account_id>/undo", methods=["POST"])
def expense_undo(account_id):
    db = get_db()
    event = db.execute(
        "select * from expense_events where account_id = ? and undone_at is null order by id desc limit 1",
        (account_id,),
    ).fetchone()
    if event is None:
        return redirect(url_for("expense_account_view", account_id=account_id))

    payload = json.loads(event["payload"])
    if event["action"] == "add":
        entry = payload.get("entry", {})
        db.execute("delete from expense_entries where id = ?", (entry.get("id"),))
    elif event["action"] == "edit":
        before = payload.get("before", {})
        db.execute(
            """update expense_entries
            set entry_date = ?, description = ?, currency = ?, direction = ?,
                amount = ?, category = ?, image_path = ?, linked_statement_entry_id = ?
            where id = ?""",
            (
                before.get("entry_date"),
                before.get("description"),
                before.get("currency"),
                before.get("direction"),
                before.get("amount"),
                before.get("category"),
                before.get("image_path"),
                before.get("linked_statement_entry_id"),
                before.get("id"),
            ),
        )
    elif event["action"] == "delete":
        entry = payload.get("entry", {})
        db.execute(
            """insert into expense_entries
            (id, account_id, seq_no, entry_date, description, currency, direction,
             amount, category, is_recurring, template_id, image_path, linked_statement_entry_id, created_at)
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                entry.get("id"),
                entry.get("account_id"),
                entry.get("seq_no", 0),
                entry.get("entry_date"),
                entry.get("description"),
                entry.get("currency"),
                entry.get("direction"),
                entry.get("amount"),
                entry.get("category"),
                entry.get("is_recurring", 0),
                entry.get("template_id"),
                entry.get("image_path"),
                entry.get("linked_statement_entry_id"),
                utc_timestamp(),
            ),
        )

    db.execute(
        "update expense_events set undone_at = ? where id = ?",
        (utc_timestamp(), event["id"]),
    )
    db.commit()
    resequence_expense_entries(account_id)
    return redirect(url_for("expense_account_view", account_id=account_id) + "#table-top")


@app.route("/expenses/accounts/<int:account_id>/templates", methods=["GET"])
def expense_templates_list(account_id):
    db = get_db()
    templates = db.execute(
        "select * from recurring_expense_templates where account_id = ? order by description",
        (account_id,),
    ).fetchall()
    return jsonify([dict(t) for t in templates])


@app.route("/expenses/accounts/<int:account_id>/templates/new", methods=["POST"])
def expense_template_new(account_id):
    db = get_db()
    account = db.execute("select id from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    db.execute(
        """insert into recurring_expense_templates
        (account_id, description, currency, direction, amount, day_of_month, category, every_n_months, is_active, last_generated, created_at)
        values (?, ?, ?, ?, ?, ?, ?, ?, 1, '', ?)""",
        (
            account_id,
            request.form["description"].strip(),
            request.form.get("currency", "CNY"),
            request.form.get("direction", "OUT"),
            float(request.form["amount"]),
            int(request.form.get("day_of_month", 1)),
            request.form.get("category", "general"),
            int(request.form.get("every_n_months", 1)),
            utc_timestamp(),
        ),
    )
    db.commit()
    return redirect(url_for("expense_account_view", account_id=account_id) + "#templates-section")


@app.route("/expenses/templates/<int:template_id>/edit", methods=["POST"])
def expense_template_edit(template_id):
    db = get_db()
    tpl = db.execute("select * from recurring_expense_templates where id = ?", (template_id,)).fetchone()
    if tpl is None:
        abort(404)
    db.execute(
        """update recurring_expense_templates
        set description = ?, currency = ?, direction = ?, amount = ?,
            day_of_month = ?, category = ?, every_n_months = ?
        where id = ?""",
        (
            request.form["description"].strip(),
            request.form.get("currency", "CNY"),
            request.form.get("direction", "OUT"),
            float(request.form["amount"]),
            int(request.form.get("day_of_month", 1)),
            request.form.get("category", "general"),
            int(request.form.get("every_n_months", 1)),
            template_id,
        ),
    )
    db.commit()
    return redirect(url_for("expense_account_view", account_id=tpl["account_id"]) + "#templates-section")


@app.route("/expenses/templates/<int:template_id>/delete", methods=["POST"])
def expense_template_delete(template_id):
    db = get_db()
    tpl = db.execute("select * from recurring_expense_templates where id = ?", (template_id,)).fetchone()
    if tpl is None:
        abort(404)
    account_id = tpl["account_id"]
    db.execute("delete from recurring_expense_templates where id = ?", (template_id,))
    db.commit()
    return redirect(url_for("expense_account_view", account_id=account_id) + "#templates-section")


@app.route("/expenses/templates/<int:template_id>/toggle", methods=["POST"])
def expense_template_toggle(template_id):
    db = get_db()
    tpl = db.execute("select * from recurring_expense_templates where id = ?", (template_id,)).fetchone()
    if tpl is None:
        abort(404)
    db.execute(
        "update recurring_expense_templates set is_active = ? where id = ?",
        (0 if tpl["is_active"] else 1, template_id),
    )
    db.commit()
    return redirect(url_for("expense_account_view", account_id=tpl["account_id"]) + "#templates-section")


@app.route("/expenses/accounts/<int:account_id>/export.csv")
def expense_export_csv(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    entries = db.execute(
        "select * from expense_entries where account_id = ? order by entry_date, id",
        (account_id,),
    ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "description", "currency", "direction", "amount", "category"])
    for entry in entries:
        writer.writerow([
            entry["entry_date"],
            entry["description"],
            entry["currency"],
            entry["direction"],
            entry["amount"],
            entry["category"],
        ])
    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = f'attachment; filename="{account["name"].replace(" ", "_").lower()}_expenses.csv"'
    return response


@app.route("/expenses/accounts/<int:account_id>/export.xlsx")
def expense_export_xlsx(account_id):
    db = get_db()
    account, currencies, rows, _ = _expense_export_data(db, account_id)
    if account is None:
        abort(404)
    return _expense_xlsx_response(account, currencies, rows)


@app.route("/expenses/accounts/<int:account_id>/export.pdf")
def expense_export_pdf(account_id):
    db = get_db()
    account, currencies, rows, totals = _expense_export_data(db, account_id)
    if account is None:
        abort(404)
    return _expense_pdf_response(account, currencies, rows, totals)


@app.route("/expenses/accounts/<int:account_id>/import", methods=["POST"])
def expense_import_csv(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        abort(404)
    upload = request.files.get("csv_file")
    if upload is None or not upload.filename:
        flash("Choose a CSV file to import", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))

    try:
        content = upload.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        flash("CSV must be UTF-8 encoded", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))

    reader = csv.DictReader(io.StringIO(content))
    rows = [dict(row) for row in reader]
    if not rows:
        flash("CSV file is empty", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))

    fieldnames = set(reader.fieldnames or [])
    if "date" not in fieldnames or "description" not in fieldnames or "amount" not in fieldnames:
        flash("CSV must include date, description, and amount columns", "error")
        return redirect(url_for("expense_account_view", account_id=account_id))

    imported = 0
    for row in rows:
        entry_date = (row.get("date") or "").strip()
        description = (row.get("description") or "").strip()
        currency = (row.get("currency") or "CNY").strip().upper()
        direction = (row.get("direction") or "OUT").strip().upper()
        amount_raw = (row.get("amount") or "").strip()
        category = (row.get("category") or "general").strip()

        if not entry_date or not description or not amount_raw:
            continue
        if currency not in ALL_EXPENSE_CURRENCIES:
            currency = "CNY"
        if direction not in ("IN", "OUT"):
            direction = "OUT"
        valid_cats = [c[0] for c in EXPENSE_CATEGORIES]
        if category not in valid_cats:
            category = "general"

        try:
            amount = float(amount_raw)
        except ValueError:
            continue

        db.execute(
            """insert into expense_entries
            (account_id, seq_no, entry_date, description, currency, direction,
             amount, category, is_recurring, template_id, created_at)
            values (?, 0, ?, ?, ?, ?, ?, ?, 0, null, ?)""",
            (account_id, entry_date, description, currency, direction, amount, category, utc_timestamp()),
        )
        imported += 1

    db.commit()
    resequence_expense_entries(account_id)
    flash(f"Imported {imported} expense entries", "success")
    return redirect(url_for("expense_account_view", account_id=account_id))


# --- Admin: API Token Management ---

@app.route("/admin/tokens")
@admin_required
def admin_tokens():
    db = get_db()
    tokens = db.execute("select * from api_tokens order by created_at desc").fetchall()
    return render_template("admin_tokens.html", tokens=tokens)


@app.route("/admin/tokens/create", methods=["POST"])
@admin_required
def admin_create_token():
    from hashlib import sha256
    name = request.form.get("name", "").strip() or "Unnamed Token"
    raw_token = f"ffs_{secrets.token_hex(32)}"
    token_hash = sha256(raw_token.encode()).hexdigest()
    token_prefix = raw_token[:12] + "..."
    db = get_db()
    db.execute(
        "insert into api_tokens(name, token_hash, token_prefix, created_at) values (?, ?, ?, ?)",
        (name, token_hash, token_prefix, utc_timestamp()),
    )
    db.commit()
    flash(f"Token created! Copy it now (shown only once): {raw_token}", "token")
    return redirect(url_for("admin_tokens"))


@app.route("/admin/tokens/<int:token_id>/revoke", methods=["POST"])
@admin_required
def admin_revoke_token(token_id):
    db = get_db()
    token = db.execute("select * from api_tokens where id = ?", (token_id,)).fetchone()
    if not token:
        abort(404)
    db.execute("update api_tokens set is_active = 0 where id = ?", (token_id,))
    db.commit()
    flash(f"Token '{token['name']}' revoked", "success")
    return redirect(url_for("admin_tokens"))


@app.route("/admin/tokens/<int:token_id>/activate", methods=["POST"])
@admin_required
def admin_activate_token(token_id):
    db = get_db()
    token = db.execute("select * from api_tokens where id = ?", (token_id,)).fetchone()
    if not token:
        abort(404)
    db.execute("update api_tokens set is_active = 1 where id = ?", (token_id,))
    db.commit()
    flash(f"Token '{token['name']}' activated", "success")
    return redirect(url_for("admin_tokens"))


@app.route("/admin/tokens/<int:token_id>/delete", methods=["POST"])
@admin_required
def admin_delete_token(token_id):
    db = get_db()
    token = db.execute("select * from api_tokens where id = ?", (token_id,)).fetchone()
    if not token:
        abort(404)
    db.execute("delete from api_tokens where id = ?", (token_id,))
    db.commit()
    flash(f"Token '{token['name']}' deleted", "success")
    return redirect(url_for("admin_tokens"))


# --- Public JSON API (token-authenticated) ---

@api_route("/api/v1/dashboard", methods=["GET"])
def api_dashboard():
    stats = dashboard_stats()
    balances = bank_balance_list()
    suppliers = supplier_balance_list()
    bank_tots = bank_balance_totals(balances)
    sup_tots = supplier_balance_totals(suppliers)
    fx = exchange_rate_summary()
    status = company_status(stats, bank_tots, sup_tots)
    return jsonify({
        "stats": stats,
        "bank_balances": balances,
        "bank_totals": bank_tots,
        "suppliers": suppliers,
        "supplier_totals": sup_tots,
        "exchange_rate_summary": fx,
        "company_status": status,
    })


@api_route("/api/v1/clients", methods=["GET"])
def api_clients():
    clients = client_list()
    groups = grouped_client_list()
    return jsonify({"clients": clients, "groups": groups})


@api_route("/api/v1/clients/<int:client_id>", methods=["GET"])
def api_client_detail(client_id):
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        return jsonify({"error": "Client not found"}), 404
    entries = db.execute(
        f"select * from statement_entries where client_id = ? order by {ENTRY_ORDER}",
        (client_id,),
    ).fetchall()
    rows = statement_rows_with_commission_state(db, client_id, entries)
    entry_dicts = []
    for r in rows:
        d = {k: r[k] for k in r}
        d["running_usd"] = r["running_usd"]
        d["running_cny"] = r["running_cny"]
        entry_dicts.append(d)
    usd_in = sum(float(e["amount"]) for e in entries if e["currency"] == "USD" and e["direction"] == "IN")
    usd_out = sum(float(e["amount"]) for e in entries if e["currency"] == "USD" and e["direction"] == "OUT")
    cny_in = sum(float(e["amount"]) for e in entries if e["currency"] == "CNY" and e["direction"] == "IN")
    cny_out = sum(float(e["amount"]) for e in entries if e["currency"] == "CNY" and e["direction"] == "OUT")
    return jsonify({
        "client": {"id": client["id"], "name": client["name"], "parent_id": client["parent_id"] if "parent_id" in client.keys() else None},
        "summary": {
            "entry_count": len(entries),
            "usd_balance": usd_in - usd_out,
            "cny_balance": cny_in - cny_out,
            "usd_in": usd_in, "usd_out": usd_out,
            "cny_in": cny_in, "cny_out": cny_out,
        },
        "entries": entry_dicts,
    })


@api_route("/api/v1/clients/<int:client_id>/export.pdf", methods=["GET"], endpoint="api_client_export_pdf")
def api_client_export_pdf(client_id):
    db = get_db()
    client, rows = _statement_export_data(db, client_id)
    if client is None:
        return jsonify({"error": "Client not found"}), 404
    try:
        return _statement_pdf_response(client, rows)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 503


@api_route("/api/v1/clients/<int:client_id>/export.xlsx", methods=["GET"], endpoint="api_client_export_xlsx")
def api_client_export_xlsx(client_id):
    db = get_db()
    client, rows = _statement_export_data(db, client_id)
    if client is None:
        return jsonify({"error": "Client not found"}), 404
    return _statement_xlsx_response(client, rows)


@api_route("/api/v1/bank-balances", methods=["GET"])
def api_bank_balances():
    balances = bank_balance_list()
    totals = bank_balance_totals(balances)
    return jsonify({"bank_balances": balances, "totals": totals})


@api_route("/api/v1/supplier-balances", methods=["GET"])
def api_supplier_balances():
    suppliers = supplier_balance_list()
    totals = supplier_balance_totals(suppliers)
    return jsonify({"suppliers": suppliers, "totals": totals})


@api_route("/api/v1/exchange-rates", methods=["GET"])
def api_exchange_rates():
    return jsonify(exchange_rate_summary())


@api_route("/api/v1/fx-rate", methods=["GET"], endpoint="api_v1_fx_rate")
def api_v1_fx_rate():
    """Get live exchange rate for any currency pair. Query params: ?from=USD&to=CNY&amount=100"""
    import time

    from_cur = request.args.get("from", "USD").upper().strip()
    to_cur = request.args.get("to", "CNY").upper().strip()
    amount = request.args.get("amount", "").strip()

    if from_cur == "RMB":
        from_cur = "CNY"
    if to_cur == "RMB":
        to_cur = "CNY"

    if from_cur == to_cur:
        ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")
        result = {"rate": 1.0, "from": from_cur, "to": to_cur, "source": "identity", "timestamp": ts}
        if amount:
            result["amount"] = float(amount)
            result["converted"] = float(amount)
        return jsonify(result)

    now = time.time()

    # Check cache
    cached = _fx_cache.get(from_cur)
    if cached and (now - cached["fetched_at"]) < _FX_CACHE_TTL:
        rate = cached["rates"].get(to_cur)
        if rate is not None:
            result = {"rate": rate, "from": from_cur, "to": to_cur, "source": cached["source"], "timestamp": cached["timestamp"]}
            if amount:
                result["amount"] = float(amount)
                result["converted"] = round(float(amount) * rate, 2)
            return jsonify(result)

    proxies = {"http": FX_PROXY_URL, "https": FX_PROXY_URL} if FX_PROXY_URL else None
    ts = datetime.now(tz=CHINA_TZ).isoformat(timespec="seconds")

    # 1) Try moneyconvert.net
    try:
        usd_rates = _fetch_moneyconvert(proxies)
        if usd_rates:
            rates = usd_rates if from_cur == "USD" else _convert_rates_from_usd(usd_rates, from_cur)
            if rates:
                _fx_cache[from_cur] = {"rates": rates, "fetched_at": now, "timestamp": ts, "source": "moneyconvert"}
                rate = rates.get(to_cur)
                if rate is not None:
                    result = {"rate": rate, "from": from_cur, "to": to_cur, "source": "moneyconvert", "timestamp": ts}
                    if amount:
                        result["amount"] = float(amount)
                        result["converted"] = round(float(amount) * rate, 2)
                    return jsonify(result)
    except Exception:
        pass

    # 2) Try open.er-api.com
    try:
        rates = _fetch_er_api(from_cur, proxies)
        if rates:
            _fx_cache[from_cur] = {"rates": rates, "fetched_at": now, "timestamp": ts, "source": "exchangerate-api"}
            rate = rates.get(to_cur)
            if rate is not None:
                result = {"rate": rate, "from": from_cur, "to": to_cur, "source": "exchangerate-api", "timestamp": ts}
                if amount:
                    result["amount"] = float(amount)
                    result["converted"] = round(float(amount) * rate, 2)
                return jsonify(result)
    except Exception:
        pass

    # 3) Fallback: stored rate from database (USD/CNY only)
    if (from_cur == "USD" and to_cur == "CNY") or (from_cur == "CNY" and to_cur == "USD"):
        try:
            summary = exchange_rate_summary()
            stored_rate = summary.get("latest_rate")
            if stored_rate is not None:
                rate = stored_rate if from_cur == "USD" else 1.0 / stored_rate
                result = {"rate": rate, "from": from_cur, "to": to_cur, "source": "stored", "timestamp": ts}
                if amount:
                    result["amount"] = float(amount)
                    result["converted"] = round(float(amount) * rate, 2)
                return jsonify(result)
        except Exception:
            pass

    return jsonify({"error": "Unable to fetch exchange rate for " + from_cur + "/" + to_cur}), 503


@api_route("/api/v1/expenses", methods=["GET"])
def api_expenses():
    accounts = expense_dashboard_data()
    return jsonify({"expense_accounts": accounts})


@api_route("/api/v1/expenses/<int:account_id>", methods=["GET"])
def api_expense_detail(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        return jsonify({"error": "Expense account not found"}), 404
    generate_recurring_expenses(account_id)
    currencies = [c.strip() for c in account["enabled_currencies"].split(",") if c.strip()]
    entries = db.execute(
        "select * from expense_entries where account_id = ? order by entry_date, id",
        (account_id,),
    ).fetchall()
    rows = expense_running_balances(entries, currencies)
    entry_dicts = []
    for r in rows:
        d = {k: r[k] for k in r if k != "running_balances"}
        d["running_balances"] = r["running_balances"]
        entry_dicts.append(d)
    summary = expense_account_summary(db, account_id)
    return jsonify({
        "account": {"id": account["id"], "name": account["name"], "currencies": currencies},
        "summary": summary,
        "entries": entry_dicts,
    })


@api_route("/api/v1/expenses/<int:account_id>/templates", methods=["GET"])
def api_expense_templates(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        return jsonify({"error": "Expense account not found"}), 404
    templates = db.execute(
        "select * from recurring_expense_templates where account_id = ? order by id",
        (account_id,),
    ).fetchall()
    return jsonify({"templates": [dict(t) for t in templates]})


@api_route("/api/v1/expenses/<int:account_id>/export.pdf", methods=["GET"], endpoint="api_expense_export_pdf")
def api_expense_export_pdf(account_id):
    db = get_db()
    account, currencies, rows, totals = _expense_export_data(db, account_id)
    if account is None:
        return jsonify({"error": "Expense account not found"}), 404
    try:
        return _expense_pdf_response(account, currencies, rows, totals)
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 503


@api_route("/api/v1/expenses/<int:account_id>/export.xlsx", methods=["GET"], endpoint="api_expense_export_xlsx")
def api_expense_export_xlsx(account_id):
    db = get_db()
    account, currencies, rows, _ = _expense_export_data(db, account_id)
    if account is None:
        return jsonify({"error": "Expense account not found"}), 404
    return _expense_xlsx_response(account, currencies, rows)


@api_route("/api/v1/users", methods=["GET"])
def api_users():
    db = get_db()
    users = db.execute("select id, username, role, is_active, last_login, created_at from users order by id").fetchall()
    return jsonify({"users": [dict(u) for u in users]})


@api_route("/api/v1/quick-submits", methods=["GET"])
def api_quick_submits():
    db = get_db()
    rows = db.execute(
        """select qs.*, c.name as client_name
        from quick_submits qs join clients c on qs.client_id = c.id
        order by qs.created_at desc"""
    ).fetchall()
    return jsonify({"quick_submits": [dict(r) for r in rows]})


@api_route("/api/v1/search", methods=["GET"])
def api_search():
    """Search entries across all clients by description keyword."""
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "Missing 'q' query parameter"}), 400
    db = get_db()
    entries = db.execute(
        """select se.*, c.name as client_name
        from statement_entries se join clients c on se.client_id = c.id
        where se.description like ?
        order by se.entry_date desc limit 100""",
        (f"%{q}%",),
    ).fetchall()
    return jsonify({"query": q, "results": [dict(e) for e in entries]})


# --- Write API endpoints (token-authenticated) ---

@api_route("/api/v1/clients", methods=["POST"], endpoint="api_create_client")
def api_create_client():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Client name is required"}), 400
    db = get_db()
    existing = db.execute("select id from clients where name = ?", (name,)).fetchone()
    if existing:
        return jsonify({"error": f"Client '{name}' already exists"}), 409
    db.execute("insert into clients(name) values (?)", (name,))
    db.commit()
    client = db.execute("select * from clients where name = ?", (name,)).fetchone()
    _api_log("create_client", "client", client["id"], {"name": name},
             {"action": "delete_client", "client_id": client["id"], "name": name})
    return jsonify({"ok": True, "client": {"id": client["id"], "name": client["name"]}}), 201


@api_route("/api/v1/clients/<int:client_id>/entries", methods=["POST"], endpoint="api_add_entry")
def api_add_entry(client_id):
    db = get_db()
    client = db.execute("select id from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        return jsonify({"error": "Client not found"}), 404
    # Support both JSON and multipart/form-data (for image uploads)
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        data = request.form.to_dict()
    else:
        data = request.get_json(silent=True) or {}
    required = ["entry_date", "description", "currency", "direction", "amount"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    currency = data["currency"].upper()
    direction = data["direction"].upper()
    if currency not in ("USD", "CNY"):
        return jsonify({"error": "Currency must be USD or CNY"}), 400
    if direction not in ("IN", "OUT"):
        return jsonify({"error": "Direction must be IN or OUT"}), 400
    kind = data.get("kind", "movement")
    if kind not in ("movement", "transfer"):
        kind = "movement"
    category_hint = data.get("category_hint", UNCATEGORIZED)
    add_to_company_profit = parse_bool_flag(data.get("add_to_company_profit"))
    try:
        profit_expense_account_id = parse_optional_int(
            data.get("profit_expense_account_id"),
            "profit_expense_account_id",
        )
        validate_profit_expense_selection(
            db,
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
            category_hint=category_hint,
            currency=currency,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    image_filename = save_upload_image(request.files.get("image"))
    db.execute(
        """insert into statement_entries
        (client_id, source_no, entry_date, description, currency, direction,
         amount, kind, category_hint, transfer_group, image_path)
        values (?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (client_id, data["entry_date"], data["description"].strip(),
         currency, direction, float(data["amount"]), kind,
         category_hint,
         data.get("transfer_group") or None, image_filename),
    )
    entry = db.execute("select * from statement_entries where client_id = ? order by id desc limit 1", (client_id,)).fetchone()
    affected_expense_accounts: set[int] = set()
    if entry is not None:
        affected_expense_accounts = sync_statement_profit_entry(
            db,
            entry["id"],
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
        )
    if data.get("transfer_group"):
        sync_exchange_group(db, data["transfer_group"])
    db.commit()
    entry = db.execute("select * from statement_entries where client_id = ? order by id desc limit 1", (client_id,)).fetchone()
    record_event(client_id, entry["id"], "add", {"entry": row_to_dict(entry)})
    resequence_client_entries(client_id)
    for expense_account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(expense_account_id)
    _api_log("create_entry", "statement_entry", entry["id"], row_to_dict(entry),
             {"action": "delete_entry", "entry_id": entry["id"], "client_id": client_id})
    return jsonify({"ok": True, "entry": row_to_dict(entry)}), 201


@api_route("/api/v1/entries/<int:entry_id>", methods=["PUT", "PATCH"], endpoint="api_update_entry")
def api_update_entry(entry_id):
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        return jsonify({"error": "Entry not found"}), 404
    before = row_to_dict(entry)
    # Support both JSON and multipart/form-data (for image uploads)
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        data = request.form.to_dict()
    else:
        data = request.get_json(silent=True) or {}
    entry_date = data.get("entry_date", entry["entry_date"])
    description = data.get("description", entry["description"]).strip()
    currency = data.get("currency", entry["currency"]).upper()
    direction = data.get("direction", entry["direction"]).upper()
    amount = float(data.get("amount", entry["amount"]))
    kind = data.get("kind", entry["kind"])
    category_hint = data.get("category_hint", entry["category_hint"])
    transfer_group = data.get("transfer_group", entry["transfer_group"]) or None
    if "add_to_company_profit" in data:
        add_to_company_profit = parse_bool_flag(data.get("add_to_company_profit"))
    else:
        add_to_company_profit = bool(entry["profit_expense_account_id"])
    if entry["commission_source_entry_id"]:
        source_entry = db.execute(
            "select currency from statement_entries where id = ?",
            (entry["commission_source_entry_id"],),
        ).fetchone()
        currency = source_entry["currency"] if source_entry else entry["currency"]
        direction = "OUT"
        kind = "movement"
        category_hint = "commission"
        transfer_group = None
    elif category_hint != "commission":
        add_to_company_profit = False
    try:
        if "profit_expense_account_id" in data:
            profit_expense_account_id = parse_optional_int(
                data.get("profit_expense_account_id"),
                "profit_expense_account_id",
            )
        else:
            profit_expense_account_id = entry["profit_expense_account_id"]
        if category_hint != "commission":
            profit_expense_account_id = None
        validate_profit_expense_selection(
            db,
            enabled=add_to_company_profit,
            account_id=profit_expense_account_id,
            category_hint=category_hint,
            currency=currency,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    old_image = entry["image_path"] if "image_path" in entry.keys() else None
    new_image = save_upload_image(request.files.get("image"))
    image_path = new_image if new_image else old_image
    if data.get("remove_image") == "1":
        image_path = None
    if old_image and old_image != image_path:
        _delete_image_file(old_image)
    db.execute(
        """update statement_entries
        set entry_date=?, description=?, currency=?, direction=?,
            amount=?, kind=?, category_hint=?, transfer_group=?, image_path=?
        where id=?""",
        (entry_date, description, currency, direction, amount, kind,
         category_hint, transfer_group, image_path, entry_id),
    )
    db.commit()
    sync_exchange_group(db, before.get("transfer_group"))
    sync_exchange_group(db, transfer_group)
    affected_expense_accounts = sync_statement_profit_entry(
        db,
        entry_id,
        enabled=add_to_company_profit,
        account_id=profit_expense_account_id,
    )
    db.commit()
    updated = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    record_event(entry["client_id"], entry_id, "edit", {"before": before, "after": row_to_dict(updated)})
    resequence_client_entries(entry["client_id"])
    for expense_account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(expense_account_id)
    _api_log("update_entry", "statement_entry", entry_id, {"before": before, "after": row_to_dict(updated)},
             {"action": "restore_entry", "entry_id": entry_id, "data": before})
    return jsonify({"ok": True, "entry": row_to_dict(updated)})


@api_route("/api/v1/entries/<int:entry_id>", methods=["DELETE"], endpoint="api_delete_entry")
def api_delete_entry(entry_id):
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        return jsonify({"error": "Entry not found"}), 404
    if commission_child_entry(db, entry_id) is not None:
        return jsonify({"error": "Delete the generated commission entry first."}), 409
    client_id = entry["client_id"]
    entry_data = row_to_dict(entry)
    record_event(client_id, entry_id, "delete", {"entry": entry_data})
    affected_expense_accounts = sync_statement_profit_entry(
        db,
        entry_id,
        enabled=False,
        account_id=None,
    )
    if entry["image_path"]:
        _delete_image_file(entry["image_path"])
    db.execute("delete from statement_entries where id = ?", (entry_id,))
    db.commit()
    resequence_client_entries(client_id)
    for expense_account_id in sorted(affected_expense_accounts):
        resequence_expense_entries(expense_account_id)
    _api_log("delete_entry", "statement_entry", entry_id, entry_data,
             {"action": "recreate_entry", "client_id": client_id, "data": entry_data})
    return jsonify({"ok": True, "deleted_id": entry_id})


@api_route("/api/v1/entries/<int:entry_id>/commission", methods=["POST"], endpoint="api_create_commission")
def api_create_commission(entry_id):
    db = get_db()
    source_entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if source_entry is None:
        return jsonify({"error": "Entry not found"}), 404
    client_id = source_entry["client_id"]
    error = commission_error_message(db, source_entry, client_id)
    if error:
        return jsonify({"error": error}), 409

    data = request.get_json(silent=True) or {}
    commission_date = (data.get("commission_date") or data.get("date") or china_today().isoformat()).strip()
    if not commission_date:
        return jsonify({"error": "commission_date is required"}), 400
    percentage_raw = data.get("percentage")
    if percentage_raw in (None, ""):
        return jsonify({"error": "percentage is required"}), 400
    note = (data.get("note") or data.get("description") or "").strip()
    try:
        created, percentage = create_commission_entry(
            db,
            source_entry,
            commission_date,
            percentage_raw,
            note,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    entry_data = row_to_dict(created)
    _api_log(
        "create_commission",
        "statement_entry",
        created["id"],
        {
            "percentage": percentage,
            "source_entry_id": source_entry["id"],
            "entry": entry_data,
        },
        {"action": "delete_entry", "entry_id": created["id"], "client_id": client_id},
    )
    return jsonify({"ok": True, "entry": entry_data, "percentage": percentage}), 201


@api_route("/api/v1/bank-balances", methods=["POST"], endpoint="api_add_bank_balance")
def api_add_bank_balance():
    data = request.get_json(silent=True) or {}
    account_name = (data.get("account_name") or "").strip()
    if not account_name:
        return jsonify({"error": "account_name is required"}), 400
    db = get_db()
    db.execute(
        "insert into bank_balances (account_name, usd_balance, cny_balance, updated_at) values (?, ?, ?, ?)",
        (account_name, float(data.get("usd_balance", 0)), float(data.get("cny_balance", 0)), utc_timestamp()),
    )
    db.commit()
    row = db.execute("select * from bank_balances order by id desc limit 1").fetchone()
    _api_log("create_bank_balance", "bank_balance", row["id"], dict(row),
             {"action": "delete_bank_balance", "id": row["id"]})
    return jsonify({"ok": True, "bank_balance": dict(row)}), 201


@api_route("/api/v1/bank-balances/<int:balance_id>", methods=["PUT", "PATCH"], endpoint="api_update_bank_balance")
def api_update_bank_balance(balance_id):
    db = get_db()
    row = db.execute("select * from bank_balances where id = ?", (balance_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Bank balance not found"}), 404
    before = dict(row)
    data = request.get_json(silent=True) or {}
    account_name = (data.get("account_name") or "").strip() or row["account_name"]
    usd = float(data.get("usd_balance", row["usd_balance"]))
    cny = float(data.get("cny_balance", row["cny_balance"]))
    db.execute(
        "update bank_balances set account_name=?, usd_balance=?, cny_balance=?, updated_at=? where id=?",
        (account_name, usd, cny, utc_timestamp(), balance_id),
    )
    db.commit()
    updated = db.execute("select * from bank_balances where id = ?", (balance_id,)).fetchone()
    _api_log("update_bank_balance", "bank_balance", balance_id, {"before": before, "after": dict(updated)},
             {"action": "restore_bank_balance", "id": balance_id, "data": before})
    return jsonify({"ok": True, "bank_balance": dict(updated)})


@api_route("/api/v1/bank-balances/<int:balance_id>", methods=["DELETE"], endpoint="api_delete_bank_balance")
def api_delete_bank_balance(balance_id):
    db = get_db()
    row = db.execute("select * from bank_balances where id = ?", (balance_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Bank balance not found"}), 404
    row_data = dict(row)
    db.execute("delete from bank_balances where id = ?", (balance_id,))
    db.commit()
    _api_log("delete_bank_balance", "bank_balance", balance_id, row_data,
             {"action": "recreate_bank_balance", "data": row_data})
    return jsonify({"ok": True, "deleted_id": balance_id})


@api_route("/api/v1/supplier-balances", methods=["POST"], endpoint="api_add_supplier")
def api_add_supplier():
    data = request.get_json(silent=True) or {}
    supplier_name = (data.get("supplier_name") or "").strip()
    if not supplier_name:
        return jsonify({"error": "supplier_name is required"}), 400
    currency = (data.get("currency") or "CNY").upper()
    if currency not in ("USD", "CNY"):
        currency = "CNY"
    db = get_db()
    db.execute(
        "insert into supplier_balances (supplier_name, currency, amount_owed, notes, updated_at) values (?, ?, ?, ?, ?)",
        (supplier_name, currency, float(data.get("amount_owed", 0)),
         (data.get("notes") or "").strip(), utc_timestamp()),
    )
    db.commit()
    row = db.execute("select * from supplier_balances order by id desc limit 1").fetchone()
    _api_log("create_supplier", "supplier_balance", row["id"], dict(row),
             {"action": "delete_supplier", "id": row["id"]})
    return jsonify({"ok": True, "supplier": dict(row)}), 201


@api_route("/api/v1/supplier-balances/<int:supplier_id>", methods=["PUT", "PATCH"], endpoint="api_update_supplier")
def api_update_supplier(supplier_id):
    db = get_db()
    row = db.execute("select * from supplier_balances where id = ?", (supplier_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Supplier not found"}), 404
    before = dict(row)
    data = request.get_json(silent=True) or {}
    supplier_name = (data.get("supplier_name") or "").strip() or row["supplier_name"]
    currency = (data.get("currency") or row["currency"]).upper()
    if currency not in ("USD", "CNY"):
        currency = row["currency"]
    amount = float(data.get("amount_owed", row["amount_owed"]))
    notes = data.get("notes", row["notes"]).strip() if data.get("notes") is not None else row["notes"]
    db.execute(
        "update supplier_balances set supplier_name=?, currency=?, amount_owed=?, notes=?, updated_at=? where id=?",
        (supplier_name, currency, amount, notes, utc_timestamp(), supplier_id),
    )
    db.commit()
    updated = db.execute("select * from supplier_balances where id = ?", (supplier_id,)).fetchone()
    _api_log("update_supplier", "supplier_balance", supplier_id, {"before": before, "after": dict(updated)},
             {"action": "restore_supplier", "id": supplier_id, "data": before})
    return jsonify({"ok": True, "supplier": dict(updated)})


@api_route("/api/v1/supplier-balances/<int:supplier_id>", methods=["DELETE"], endpoint="api_delete_supplier")
def api_delete_supplier(supplier_id):
    db = get_db()
    row = db.execute("select * from supplier_balances where id = ?", (supplier_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Supplier not found"}), 404
    row_data = dict(row)
    db.execute("delete from supplier_balances where id = ?", (supplier_id,))
    db.commit()
    _api_log("delete_supplier", "supplier_balance", supplier_id, row_data,
             {"action": "recreate_supplier", "data": row_data})
    return jsonify({"ok": True, "deleted_id": supplier_id})


@api_route("/api/v1/expenses/<int:account_id>/entries", methods=["POST"], endpoint="api_add_expense_entry")
def api_add_expense_entry(account_id):
    db = get_db()
    account = db.execute("select * from expense_accounts where id = ?", (account_id,)).fetchone()
    if account is None:
        return jsonify({"error": "Expense account not found"}), 404
    data = request.get_json(silent=True) or {}
    required = ["entry_date", "description", "amount"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400
    currency = (data.get("currency") or "CNY").upper()
    if currency not in ALL_EXPENSE_CURRENCIES:
        currency = "CNY"
    direction = (data.get("direction") or "OUT").upper()
    if direction not in ("IN", "OUT"):
        direction = "OUT"
    category = data.get("category", "general")
    valid_cats = [c[0] for c in EXPENSE_CATEGORIES]
    if category not in valid_cats:
        category = "general"
    db.execute(
        """insert into expense_entries
        (account_id, seq_no, entry_date, description, currency, direction,
         amount, category, is_recurring, template_id, created_at)
        values (?, 0, ?, ?, ?, ?, ?, ?, 0, null, ?)""",
        (account_id, data["entry_date"], data["description"].strip(),
         currency, direction, float(data["amount"]), category, utc_timestamp()),
    )
    db.commit()
    entry = db.execute("select * from expense_entries where account_id = ? order by id desc limit 1", (account_id,)).fetchone()
    expense_record_event(account_id, entry["id"], "add", {"entry": expense_entry_to_dict(entry)})
    resequence_expense_entries(account_id)
    _api_log("create_expense_entry", "expense_entry", entry["id"], expense_entry_to_dict(entry),
             {"action": "delete_expense_entry", "entry_id": entry["id"], "account_id": account_id})
    return jsonify({"ok": True, "entry": expense_entry_to_dict(entry)}), 201


@api_route("/api/v1/clients/<int:client_id>/exchange", methods=["POST"], endpoint="api_exchange")
def api_exchange(client_id):
    """Create a USD-to-CNY exchange pair for a client."""
    db = get_db()
    client = db.execute("select * from clients where id = ?", (client_id,)).fetchone()
    if client is None:
        return jsonify({"error": "Client not found"}), 404
    data = request.get_json(silent=True) or {}
    exchange_date = data.get("exchange_date", china_today().isoformat())
    usd_amount = float(data.get("usd_amount", 0))
    cny_amount = float(data.get("cny_amount", 0))
    exchange_rate = float(data["exchange_rate"]) if data.get("exchange_rate") else None
    note = (data.get("note") or "USD to RMB exchange").strip()
    if usd_amount <= 0:
        return jsonify({"error": "usd_amount must be positive"}), 400
    if exchange_rate and cny_amount <= 0:
        cny_amount = round(usd_amount * exchange_rate, 2)
    if cny_amount <= 0:
        return jsonify({"error": "cny_amount or exchange_rate required"}), 400
    effective_rate = cny_amount / usd_amount
    transfer_group = make_transfer_group()
    for cur, dirn, amt, desc in [("USD", "OUT", usd_amount, note), ("CNY", "IN", cny_amount, note)]:
        db.execute(
            """insert into statement_entries
            (client_id, source_no, entry_date, description, currency, direction,
             amount, kind, category_hint, transfer_group, exchange_rate)
            values (?, 0, ?, ?, ?, ?, ?, 'transfer', 'fx_transfer', ?, ?)""",
            (client_id, exchange_date, desc, cur, dirn, amt, transfer_group, effective_rate),
        )
    sync_exchange_group(db, transfer_group)
    db.commit()
    resequence_client_entries(client_id)
    entries = db.execute(
        "select * from statement_entries where transfer_group = ? order by id",
        (transfer_group,),
    ).fetchall()
    entry_ids = [e["id"] for e in entries]
    _api_log("create_exchange", "statement_entry", entry_ids[0] if entry_ids else None,
             {"transfer_group": transfer_group, "entries": [row_to_dict(e) for e in entries]},
             {"action": "delete_exchange_entries", "entry_ids": entry_ids, "client_id": client_id})
    return jsonify({"ok": True, "transfer_group": transfer_group, "entries": [row_to_dict(e) for e in entries]}), 201


# --- Audit Log & Undo ---

@api_route("/api/v1/audit-log", methods=["GET"], endpoint="api_audit_log")
def api_audit_log():
    db = get_db()
    action_filter = request.args.get("action", "").strip()
    limit = min(int(request.args.get("limit", 50)), 500)
    offset = int(request.args.get("offset", 0))
    if action_filter:
        rows = db.execute(
            "select * from api_audit_log where action = ? order by id desc limit ? offset ?",
            (action_filter, limit, offset),
        ).fetchall()
    else:
        rows = db.execute(
            "select * from api_audit_log order by id desc limit ? offset ?",
            (limit, offset),
        ).fetchall()
    total = db.execute("select count(*) from api_audit_log").fetchone()[0]
    results = []
    for r in rows:
        d = dict(r)
        d["detail"] = json.loads(d["detail"]) if d["detail"] else {}
        d["undo_data"] = json.loads(d["undo_data"]) if d["undo_data"] else None
        results.append(d)
    return jsonify({"audit_log": results, "total": total, "limit": limit, "offset": offset})


@api_route("/api/v1/undo", methods=["POST"], endpoint="api_undo")
def api_undo():
    """Undo the most recent un-undone API write action."""
    db = get_db()
    log_entry = db.execute(
        "select * from api_audit_log where undone = 0 and undo_data is not null order by id desc limit 1"
    ).fetchone()
    if not log_entry:
        return jsonify({"error": "Nothing to undo"}), 404
    undo = json.loads(log_entry["undo_data"])
    action = undo.get("action", "")
    result_detail = {"undone_log_id": log_entry["id"], "original_action": log_entry["action"]}

    try:
        if action == "delete_client":
            db.execute("delete from statement_entries where client_id = ?", (undo["client_id"],))
            db.execute("delete from clients where id = ?", (undo["client_id"],))
            result_detail["deleted_client_id"] = undo["client_id"]

        elif action == "delete_entry":
            entry_id = undo["entry_id"]
            affected_expense_accounts = sync_statement_profit_entry(
                db,
                entry_id,
                enabled=False,
                account_id=None,
            )
            db.execute("delete from statement_entries where id = ?", (entry_id,))
            resequence_client_entries(undo["client_id"])
            for account_id in sorted(affected_expense_accounts):
                resequence_expense_entries(account_id)
            result_detail["deleted_entry_id"] = entry_id

        elif action == "restore_entry":
            d = undo["data"]
            db.execute(
                """update statement_entries
                set entry_date=?, description=?, currency=?, direction=?,
                    amount=?, kind=?, category_hint=?, transfer_group=?, image_path=?,
                    commission_source_entry_id=?
                where id=?""",
                (d["entry_date"], d["description"], d["currency"], d["direction"],
                 d["amount"], d["kind"], d["category_hint"], d.get("transfer_group"),
                 d.get("image_path"), d.get("commission_source_entry_id"), undo["entry_id"]),
            )
            resequence_client_entries(d["client_id"])
            affected_expense_accounts = sync_statement_profit_entry(
                db,
                undo["entry_id"],
                enabled=bool(d.get("profit_expense_account_id")),
                account_id=d.get("profit_expense_account_id"),
            )
            for account_id in sorted(affected_expense_accounts):
                resequence_expense_entries(account_id)
            result_detail["restored_entry_id"] = undo["entry_id"]

        elif action == "recreate_entry":
            d = undo["data"]
            cursor = db.execute(
                """insert into statement_entries
                (client_id, source_no, entry_date, description, currency, direction,
                 amount, kind, category_hint, transfer_group, exchange_rate, image_path,
                 commission_source_entry_id, profit_expense_entry_id, profit_expense_account_id)
                values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (d["client_id"], d["source_no"], d["entry_date"], d["description"],
                 d["currency"], d["direction"], d["amount"], d["kind"], d["category_hint"],
                 d.get("transfer_group"), d.get("exchange_rate"), d.get("image_path"),
                 d.get("commission_source_entry_id"), d.get("profit_expense_entry_id"),
                 d.get("profit_expense_account_id")),
            )
            resequence_client_entries(undo["client_id"])
            recreated_id = cursor.lastrowid
            affected_expense_accounts = sync_statement_profit_entry(
                db,
                recreated_id,
                enabled=bool(d.get("profit_expense_account_id")),
                account_id=d.get("profit_expense_account_id"),
            )
            for account_id in sorted(affected_expense_accounts):
                resequence_expense_entries(account_id)
            result_detail["recreated_for_client"] = undo["client_id"]

        elif action == "delete_bank_balance":
            db.execute("delete from bank_balances where id = ?", (undo["id"],))
            result_detail["deleted_bank_balance_id"] = undo["id"]

        elif action == "restore_bank_balance":
            d = undo["data"]
            db.execute(
                "update bank_balances set account_name=?, usd_balance=?, cny_balance=?, updated_at=? where id=?",
                (d["account_name"], d["usd_balance"], d["cny_balance"], d["updated_at"], undo["id"]),
            )
            result_detail["restored_bank_balance_id"] = undo["id"]

        elif action == "recreate_bank_balance":
            d = undo["data"]
            db.execute(
                "insert into bank_balances (account_name, usd_balance, cny_balance, updated_at) values (?, ?, ?, ?)",
                (d["account_name"], d["usd_balance"], d["cny_balance"], d["updated_at"]),
            )
            result_detail["recreated_bank_balance"] = d["account_name"]

        elif action == "delete_supplier":
            db.execute("delete from supplier_balances where id = ?", (undo["id"],))
            result_detail["deleted_supplier_id"] = undo["id"]

        elif action == "restore_supplier":
            d = undo["data"]
            db.execute(
                "update supplier_balances set supplier_name=?, currency=?, amount_owed=?, notes=?, updated_at=? where id=?",
                (d["supplier_name"], d["currency"], d["amount_owed"], d["notes"], d["updated_at"], undo["id"]),
            )
            result_detail["restored_supplier_id"] = undo["id"]

        elif action == "recreate_supplier":
            d = undo["data"]
            db.execute(
                "insert into supplier_balances (supplier_name, currency, amount_owed, notes, updated_at) values (?, ?, ?, ?, ?)",
                (d["supplier_name"], d["currency"], d["amount_owed"], d["notes"], d["updated_at"]),
            )
            result_detail["recreated_supplier"] = d["supplier_name"]

        elif action == "delete_expense_entry":
            db.execute("delete from expense_entries where id = ?", (undo["entry_id"],))
            resequence_expense_entries(undo["account_id"])
            result_detail["deleted_expense_entry_id"] = undo["entry_id"]

        elif action == "delete_exchange_entries":
            for eid in undo.get("entry_ids", []):
                db.execute("delete from statement_entries where id = ?", (eid,))
            resequence_client_entries(undo["client_id"])
            result_detail["deleted_exchange_entry_ids"] = undo["entry_ids"]

        else:
            return jsonify({"error": f"Unknown undo action: {action}"}), 400

        db.execute("update api_audit_log set undone = 1 where id = ?", (log_entry["id"],))
        db.commit()

    except Exception as exc:
        db.rollback()
        return jsonify({"error": f"Undo failed: {exc}"}), 500

    return jsonify({"ok": True, "undone": result_detail})


# --- API: Image upload/delete for statement entries ---

@api_route("/api/v1/entries/<int:entry_id>/image", methods=["POST"], endpoint="api_upload_entry_image")
def api_upload_entry_image(entry_id):
    """Upload or replace an image for an existing statement entry."""
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        return jsonify({"error": "Entry not found"}), 404
    image_file = request.files.get("image")
    if not image_file or not image_file.filename:
        return jsonify({"error": "No image file provided. Send as multipart/form-data with field name 'image'"}), 400
    new_image = save_upload_image(image_file)
    if not new_image:
        return jsonify({"error": f"Invalid image type. Allowed: {', '.join(ALLOWED_IMAGE_EXT)}"}), 400
    old_image = entry["image_path"] if "image_path" in entry.keys() else None
    if old_image:
        _delete_image_file(old_image)
    db.execute("update statement_entries set image_path = ? where id = ?", (new_image, entry_id))
    db.commit()
    updated = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    record_event(entry["client_id"], entry_id, "edit", {"before": row_to_dict(entry), "after": row_to_dict(updated)})
    _api_log("upload_image", "statement_entry", entry_id,
             {"old_image": old_image, "new_image": new_image},
             {"action": "restore_entry", "entry_id": entry_id, "data": row_to_dict(entry)})
    return jsonify({"ok": True, "entry": row_to_dict(updated)})


@api_route("/api/v1/entries/<int:entry_id>/image", methods=["DELETE"], endpoint="api_delete_entry_image")
def api_delete_entry_image(entry_id):
    """Remove the image from a statement entry."""
    db = get_db()
    entry = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    if entry is None:
        return jsonify({"error": "Entry not found"}), 404
    old_image = entry["image_path"] if "image_path" in entry.keys() else None
    if not old_image:
        return jsonify({"error": "Entry has no image"}), 404
    _delete_image_file(old_image)
    db.execute("update statement_entries set image_path = null where id = ?", (entry_id,))
    db.commit()
    updated = db.execute("select * from statement_entries where id = ?", (entry_id,)).fetchone()
    record_event(entry["client_id"], entry_id, "edit", {"before": row_to_dict(entry), "after": row_to_dict(updated)})
    _api_log("delete_image", "statement_entry", entry_id,
             {"deleted_image": old_image},
             {"action": "restore_entry", "entry_id": entry_id, "data": row_to_dict(entry)})
    return jsonify({"ok": True, "entry": row_to_dict(updated)})


if __name__ == "__main__":
    init_db()
    ensure_seeded()
    resequence_all_clients()
    app.run(
        host=os.environ.get("HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5050")),
        debug=os.environ.get("FLASK_DEBUG", "0") == "1",
    )
